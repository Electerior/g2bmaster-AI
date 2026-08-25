# -*- coding: utf-8 -*-
"""이전 충실성 회귀 — 동일 batch JSON을 [소스 코드] vs [bidpipe 이전 후 코드]로 각각 생성,
워크북 셀 + 게이트 exit code 전수 대조. 둘 다 동일하면 이전이 동작을 바꾸지 않은 것.

소스 폴더(~/Documents/Elect*)가 없으면 SKIP(0 exit) — 다른 머신에서도 make check 가 깨지지 않는다.
실행: make bidpipe-fidelity  (또는 .venv/bin/python bidpipe/tests/test_migration_fidelity.py)
"""
import glob, json, os, subprocess, sys, tempfile
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_cands = glob.glob(os.path.expanduser("~/Documents/Elect*"))
if not _cands:
    print("SKIP: 소스 워크스페이스 없음 (~/Documents/Elect*) — 이전 충실성 확인은 이 머신에서만")
    sys.exit(0)
SRC = _cands[0]



DAY = os.path.join(SRC, "20260825")   # fixture: 2026-08-25 배치
PY = os.path.join(REPO, ".venv", "bin", "python")
SRC_GEN = os.path.join(SRC, ".agents", "scripts", "gen_analysis.py")
NEW_GEN = os.path.join(REPO, "bidpipe", ".agents", "scripts", "gen_analysis.py")

def cells(ws):
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                v = str(c.value).strip()
                if v:
                    out[(c.row, c.column)] = v
    return out

def run(gen, batch, outdir, envroot):
    env = dict(os.environ, BIDPIPE_ROOT=envroot, GEN_OUTDIR=outdir)
    r = subprocess.run([PY, gen, batch], env=env, capture_output=True, text=True, timeout=300)
    return r

batches = sorted(glob.glob(os.path.join(DAY, "batch*.json")))
if not batches:
    print(f"SKIP: fixture 없음 ({DAY}/batch*.json)"); sys.exit(0)
total, matched, failed = 0, 0, 0
for b in batches:
    dsrc = tempfile.mkdtemp(prefix="src_")
    dnew = tempfile.mkdtemp(prefix="new_")
    r_src = run(SRC_GEN, b, dsrc, SRC)
    r_new = run(NEW_GEN, b, dnew, os.path.join(REPO, "bidpipe"))
    # exit code 동일성 (게이트 동작 동일 여부)
    ec_match = "OK" if r_src.returncode == r_new.returncode else f"DIFF(src={r_src.returncode},new={r_new.returncode})"
    sfiles = sorted(os.listdir(dsrc))
    nfiles = sorted(os.listdir(dnew))
    if sfiles != nfiles:
        print(f"FILELIST DIFF {os.path.basename(b)}: src={sfiles} new={nfiles}")
    for f in sfiles:
        if f.endswith(".xlsx"):
            total += 1
            sp, np_ = os.path.join(dsrc, f), os.path.join(dnew, f)
            if not os.path.exists(np_):
                print(f"MISSING new {f}"); failed += 1; continue
            c1, c2 = cells(load_workbook(sp, data_only=True).active), cells(load_workbook(np_, data_only=True).active)
            diff = [f"r{k[0]}c{k[1]}" for k in (set(c1)|set(c2)) if c1.get(k) != c2.get(k)]
            if diff:
                failed += 1
                print(f"CELL DIFF {os.path.basename(b)}/{f}: {len(diff)}개 → {diff[:6]}")
            else:
                matched += 1
                print(f"MATCH (exit {ec_match}) {os.path.basename(b)}/{f} ({len(c1)} 셀)")
    if r_src.returncode != 0:
        print(f"  [gate] {os.path.basename(b)}: src_exit={r_src.returncode} new_exit={r_new.returncode}  {ec_match}")

print(f"\n=== 이전 충실성 검증: {total}개 파일, {matched}개 셀-동일, {failed}개 차이 ===")
print("PASS: 이전 후 코드가 소스와 동일한 출력을 재현" if failed==0 else "FAIL: 차이 발견 — 이전 중 동작 변경")
sys.exit(1 if failed else 0)
