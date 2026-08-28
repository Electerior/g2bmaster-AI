# -*- coding: utf-8 -*-
"""날짜 폴더의 분석 워크북을 훑어 대시보드용 JSON을 만든다.

날짜 폴더(YYYYMMDD/)가 사실상 DB다. 이 스크립트는 그 DB의 인덱스를 뽑는다.

    python .agents/scripts/scan_dashboard.py            # stdout으로 JSON
    python .agents/scripts/scan_dashboard.py -o out.json

마진 계산·추정가 감사는 audit_prices를 **그대로 import**해서 쓴다.
대시보드가 별도 계산식을 갖게 되면 감사 규칙이 바뀔 때 조용히 어긋나기 때문이다.

파싱이 실패한 파일도 결과에서 빼지 않고 status="unreadable"로 실어 보낸다.
목록에서 조용히 사라지는 공고가 있으면 그게 제일 위험하다(AGENTS.md 규칙 2의 정신).
"""
import sys, os, glob, json, re, argparse, datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))       # .../Electerior
sys.path.insert(0, HERE)

from openpyxl import load_workbook
import audit_prices
from audit_prices import EST_SHARE_BLOCK
import dashboard_view

# 분석 시트 메타 셀 (gen_analysis.py가 쓰는 고정 좌표)
META = {
    "title":     "B4",   "notice_no": "B5",  "agency":    "F5",
    "posted_at": "B6",   "deadline":  "F6",
    "open_at":   "B7",   "delivery":  "F7",
    "method":    "B8",   "award":     "F8",
    "eligibility": "B9",
    "place":     "B10",  "url":       "F10",
    "est_price": "B13",  "base":      "B14",  "rate": "B15",
    "summary":   "A19",
}


def _s(v):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d %H:%M") if isinstance(v, dt.datetime) else v.strftime("%Y-%m-%d")
    return str(v).strip()


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _deadline_iso(raw):
    """'2026-08-27 10:00' / '2026-08-27' / '2026.08.27 10시' 등을 ISO로."""
    if not raw:
        return None
    m = re.search(r"(\d{4})[-./년\s]+(\d{1,2})[-./월\s]+(\d{1,2})", raw)
    if not m:
        return None
    y, mo, d = (int(x) for x in m.groups())
    hm = re.search(r"(\d{1,2})\s*[:시]\s*(\d{2})?", raw[m.end():])
    hh, mi = (int(hm.group(1)), int(hm.group(2) or 0)) if hm else (0, 0)
    try:
        return dt.datetime(y, mo, d, min(hh, 23), min(mi, 59)).isoformat()
    except ValueError:
        return None


def _floor_rate_note(award, rate):
    """이율의 근거가 공고 명시인지 기본값인지 구분 (AGENTS.md 규칙 7)."""
    if re.search(r"낙찰\s*하한|하한율|적격심사", award or ""):
        return "공고명시"
    if rate in (0.96, 0.88):
        return "기본값"
    return "확인필요"


def scan_file(path):
    rel = os.path.relpath(path, ROOT)
    folder = rel.split(os.sep)[0]
    rec = {
        "file": rel, "folder": folder, "filename": os.path.basename(path),
        "mtime": dt.datetime.fromtimestamp(os.path.getmtime(path)).isoformat(timespec="seconds"),
        "status": "ok",
    }
    try:
        wb = load_workbook(path, data_only=False)
        if "분석" not in wb.sheetnames:
            rec["status"] = "unreadable"
            rec["error"] = "'분석' 시트 없음"
            return rec
        ws = wb["분석"]
        for key, cell in META.items():
            rec[key] = _s(ws[cell].value)
        rec["est_price"] = _num(ws[META["est_price"]].value)
        rec["base"] = _num(ws[META["base"]].value)
        rec["rate"] = _num(ws[META["rate"]].value)
        rec["summary"] = (rec.get("summary") or "")[:600]
        rec["deadline_iso"] = _deadline_iso(rec.get("deadline"))
        rec["rate_basis"] = _floor_rate_note(rec.get("award"), rec.get("rate"))
        # 선택 행 수 / 미달·확인필요 행 수 (규칙 8)
        end, _ = audit_prices._layout(ws)
        sel = unmet = 0
        for r in range(audit_prices.R0, end + 1):
            if not ws.cell(row=r, column=1).value:
                continue
            if ws.cell(row=r, column=7).value == "O":
                sel += 1
                note = str(ws.cell(row=r, column=9).value or "")
                if ("미달" in note) or ("확인필요" in note):
                    unmet += 1
        rec["sel_rows"], rec["unmet_rows"] = sel, unmet
    except Exception as e:
        rec["status"] = "unreadable"
        rec["error"] = f"{type(e).__name__}: {e}"
        return rec

    # 마진·추정가 감사 (audit_prices 재사용)
    try:
        a = audit_prices.audit_file(path)
    except Exception as e:
        a = None
        rec["audit_error"] = f"{type(e).__name__}: {e}"
    if a:
        rec.update({
            "bid": a["bid"], "cost": a["cost"], "margin": a["margin"],
            "margin_rate": a["margin_rate"], "est_rows": a["est_rows"],
            "est_amount": a["est_amount"], "est_share": a["est_share"],
            "zero_quote_rows": a["zero_quote_rows"], "bypass_rows": a["bypass_rows"],
            "reportable": a["reportable"],
            # 규칙 15 — 원가가 아니다. 화면·시트에 '참고'로만 싣는다
            "extra_cost": a.get("extra_cost", 0),
        })
        # 원가 0원짜리는 마진율이 100%로 찍힌다. 물품행이 아예 없거나(참가불가 검토건)
        # 견적 행을 전부 0원으로 넣은 경우다 — 둘 다 '마진 100%'가 아니라 '원가 미입력'이다.
        # 그대로 정렬에 태우면 목록 맨 위를 가짜 100%가 점거한다 (AGENTS.md 규칙 6).
        if not a["cost"]:
            rec["status"] = "no_cost"
            rec["margin_rate"] = rec["margin"] = None
            rec["reportable"] = False
    else:
        rec["status"] = "no_margin" if rec["status"] == "ok" else rec["status"]
        rec.update({"bid": None, "cost": None, "margin": None, "margin_rate": None,
                    "est_share": None, "reportable": False})
    return rec


FNAME_RE = re.compile(r"^\d{8}_([A-Za-z0-9]+)_")


def _looks_like_notice_no(tag):
    """파일명 2번째 토큰이 공고번호인지 기관명 약자인지 가른다.
    `20260819_KETI_Q30_...`처럼 기관명을 쓰는 관례가 있어서, 숫자 없는 순수 알파벳은
    불일치로 세지 않는다 (안 그러면 KETI/ETRI 7건이 전부 오탄으로 뜼고,
    진짜 사고 1건이 그 속에 묻힌다)."""
    return bool(tag) and len(tag) >= 6 and any(c.isdigit() for c in tag)


def cross_check(rows):
    """목록 단위로만 보이는 사고를 잡는다.

    같은 공고를 두 번 분석하면 마진율 표에 같은 건이 두 줄로 앉고, 파일명과 내용이
    어긋나면 '분석했다고 믿는 공고'가 실제로는 없다.
    (2026-08-21 실제 발견: 20260819/20260818_R26DD20847739_대구대.xlsx는
     동남보건대 파일의 바이트 단위 복사본이었다 — 대구대 R26DD20847739 분석은 존재하지 않았다)
    """
    by_no = {}
    for r in rows:
        no = (r.get("notice_no") or "").strip()
        if no:
            by_no.setdefault(no, []).append(r)

    import hashlib
    digests = {}
    for r in rows:
        try:
            with open(os.path.join(ROOT, r["file"]), "rb") as fp:
                digests.setdefault(hashlib.md5(fp.read()).hexdigest(), []).append(r["file"])
        except OSError:
            pass

    dup_files = {f: g for g in digests.values() if len(g) > 1 for f in g}

    for r in rows:
        no = (r.get("notice_no") or "").strip()
        siblings = [x["file"] for x in by_no.get(no, []) if x["file"] != r["file"]]
        r["dup_notice"] = siblings
        m = FNAME_RE.match(r["filename"])
        tag = m.group(1) if m else None
        r["name_mismatch"] = bool(
            _looks_like_notice_no(tag) and no and not no.upper().startswith(tag.upper())
        )
        r["name_tag"] = tag
        twins = [f for f in dup_files.get(r["file"], []) if f != r["file"]]
        r["identical_to"] = twins
    return rows


def attach_view(rows):
    """표기 규칙(규칙 6)을 여기서 굳혀 내보낸다.

    프론트(web/format.ts)도 시트 내보내기(sheet_export.py)도 이 값을 **표시만** 한다.
    표기 로직이 소비자마다 따로 있으면 규칙이 바뀔 때 한쪽만 고쳐져 갈라진다.
    무결성 라벨도 같은 이유로 여기서 만든다.
    """
    for r in rows:
        r["view"] = dashboard_view.margin_view(r, EST_SHARE_BLOCK)
        r["integrity"] = dashboard_view.integrity_label(r)
    return rows


MIRROR_CFG = os.path.join(ROOT, ".agents", "data", "sheet_mirror.json")


def mirror_config():
    """구글시트 미러 설정. 대시보드가 iframe으로 박을 때 쓴다.

    번들에 doc id를 박지 않기 위해 서버 응답으로 내려보낸다.
    설정이 없으면 None — 프론트가 '미설정' 안내를 띄운다.
    """
    try:
        with open(MIRROR_CFG, encoding="utf-8") as fp:
            cfg = json.load(fp)
    except (OSError, ValueError):
        return None
    doc = cfg.get("doc_id")
    if not doc:
        return None
    return {
        "doc_id": doc,
        "url": cfg.get("url") or f"https://docs.google.com/spreadsheets/d/{doc}/edit",
        "account": cfg.get("account"),
        # [{name, gid}] — 탭 전환용
        "sheets": [{"name": k, "gid": str(v.get("gid", "0"))}
                   for k, v in (cfg.get("sheets") or {}).items()],
    }


def build_payload(pattern="20*/*.xlsx"):
    """스캔 결과 전체. 대시보드 서버와 sheet_export.py가 같이 쓴다."""
    files = sorted(glob.glob(os.path.join(ROOT, pattern)))
    rows = attach_view(cross_check([scan_file(f) for f in files]))
    # 정렬도 view["sort"]로 한다. margin_rate 로 직접 줄을 세우면 '판정보류'로 가린 건이
    # 실제 마진율 순서대로 앉아, 수치를 가려놓고 위치로 알려주는 꼴이 된다 (규칙 6).
    rows.sort(key=lambda r: (r["view"]["sort"] is None, -(r["view"]["sort"] or 0)))

    return {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "root": ROOT,
        "est_share_block": EST_SHARE_BLOCK,
        "count": len(rows),
        "issues": {
            "dup_notice": sum(1 for r in rows if r.get("dup_notice")),
            "name_mismatch": sum(1 for r in rows if r.get("name_mismatch")),
            "identical": sum(1 for r in rows if r.get("identical_to")),
        },
        "folders": sorted({r["folder"] for r in rows}, reverse=True),
        "mirror": mirror_config(),
        "rows": rows,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--out")
    ap.add_argument("--glob", default="20*/*.xlsx")
    args = ap.parse_args()

    out = build_payload(args.glob)
    rows = out["rows"]
    text = json.dumps(out, ensure_ascii=False, indent=1)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as fp:
            fp.write(text)
        bad = sum(1 for r in rows if r["status"] != "ok")
        print(f"{len(rows)}건 → {args.out} (문제 {bad}건)", file=sys.stderr)
    else:
        print(text)


if __name__ == "__main__":
    main()
