# -*- coding: utf-8 -*-
"""가격 무인 갱신기 — LLM 호출 없이 시세를 다시 읽고, 마진에 영향을 준 것만 골라낸다.

설계 원칙
  상품이 특정된 뒤의 가격 갱신은 판단이 아니라 결정론적 스크레이핑이다.
  따라서 갱신은 무제한 반복해도 비용이 0이고, AI는 '예외'만 본다.

  1층  갱신   : 살아있는 공고(개찰 전)에 걸린 상품 + 재등장 코어 상품만 긁는다.
                개찰이 지난 공고의 상품은 자동 은퇴 → 대상이 무한히 늘지 않는다.
  2층  게이트 : 단가 변동률이 아니라 '그 공고의 마진율을 몇 %p 흔드는가'로 거른다.
                임계 미만은 로그만 쌓고 끝. 임계 이상 + 스크레이핑 실패만 AI에게 넘긴다.

사용법
  python price_refresh.py                      # 갱신 + 예외 리포트
  python price_refresh.py --threshold 1.0      # 임계 1.0%p (기본 0.5)
  python price_refresh.py --all                # 개찰 지난 공고까지 포함
  python price_refresh.py --dry                # 관측로그 기록 없이 조회만

산출물
  .agents/data/price/observations.jsonl   append-only 관측 로그
  .agents/data/price/exceptions.json      AI가 봐야 할 건만
"""
import os, re, io, sys, json, glob, time, random, datetime, urllib.request, urllib.error
from urllib.parse import urlparse, parse_qs
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# 스키마는 price_schema가 단일 기준이다 (상품키·티어·관측 레코드·파생 뷰).
from price_schema import (ROOT, DATA, OBS, product_key, observe, append_obs,
                          latest_by_key, load_products, save_products,
                          rebuild_view, extract_model_code)

EXC = os.path.join(DATA, "exceptions.json")
os.makedirs(DATA, exist_ok=True)

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
DATE_PAT = re.compile(r"(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})")
MD_PAT = re.compile(r"(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)")

# ── 스크레이퍼 ────────────────────────────────────────────
def _safe_url(u):
    """한글이 들어있는 URL은 percent-encoding 해야 요청이 간다 (Dell 한국몰 등)."""
    try:
        u.encode("ascii")
        return u
    except UnicodeEncodeError:
        p = urlparse(u)
        return p._replace(path=urllib.parse.quote(p.path, safe="/%"),
                          query=urllib.parse.quote(p.query, safe="=&%")).geturl()


def _get(url, timeout=9):
    req = urllib.request.Request(_safe_url(url), headers={
        "User-Agent": UA, "Accept-Language": "ko-KR,ko;q=0.9",
        "Accept": "text/html,application/xhtml+xml"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    for enc in ("utf-8", "euc-kr", "cp949"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "ignore")


def page_title(html):
    """상품 페이지 제목. 워크북 라벨↔URL 불일치 자동 검출에 쓴다."""
    m = re.search(r"<title>(.*?)</title>", html, re.S)
    if not m:
        return None
    t = re.sub(r"\s+", " ", m.group(1)).strip()
    for tail in (" : 다나와 가격비교", " : 다나와 통합검색", " - 11번가", " : 롯데ON"):
        if t.endswith(tail):
            t = t[: -len(tail)].strip()
    return t or None


def scrape(url, _retry=2):
    """{price, vendor, delivery, method, title} 또는 {error}. LLM 없음."""
    try:
        html = _get(url)
    except urllib.error.HTTPError as e:
        # 403/429는 봇 차단이지 끊어진 링크가 아니다 (쿠팡·G마켓·스마트스토어)
        if e.code in (429, 503) and _retry:
            time.sleep(2.5 * (3 - _retry) + random.uniform(0, 1.5))
            return scrape(url, _retry - 1)
        return {"error": "HTTP %s" % e.code, "blocked": e.code in (403, 429, 503)}
    except Exception as e:
        # 다나와는 동시요청에 레이트리밋으로 URLError를 던진다. 1회 재시도로는 부족했다
        # (2026-08-21 실측: 230개 중 다나와 37건이 URLError로 새어 커버리지가 60%까지 떨어졌다).
        # 지수 백오프 + 지터로 2회까지 물러선다.
        if _retry:
            time.sleep(1.5 * (3 - _retry) + random.uniform(0, 1.5))
            return scrape(url, _retry - 1)
        return {"error": type(e).__name__}

    # 본문이 터무니없이 짧으면 JS 리다이렉트·봇벽이다 (컴퓨존은 83바이트만 준다)
    if len(html) < 800:
        return {"error": "빈 응답(JS 리다이렉트)", "blocked": True}

    title = page_title(html)

    # ── 즉시할인이 fetch에 안 잡히는 몰 → 브라우저 티어로 넘긴다 ──
    # (2026-08-21) 파서 버그가 아니라 **응답 자체가 정가**다. 여기서 정가를 받아들이면
    # 있지도 않은 급등이 만들어진다(롯데온 다크플래시 팬 78,050원을 141,900원으로 읽어 +81%).
    if "lotteon.com" in url:
        # 롯데온 JSON-LD의 offers.price는 '할인전 가격'이다. 실판매가는 JS 렌더 후에만 나온다.
        return {"error": "정가만 노출(할인 미반영)", "blocked": True, "title": title}
    if "11st.co.kr" in url:
        o = re.search(r'"offers"\s*:\s*\{[^{}]*"price"\s*:\s*"?([\d.]+)', html)
        s = re.search(r'"priceSpecification"\s*:\s*\{[^{}]*"price"\s*:\s*"?([\d.]+)'
                      r'[^{}]*StrikethroughPrice', html)
        if o:
            price, strike = int(float(o.group(1))), (int(float(s.group(1))) if s else None)
            # 정상 응답이면 offers.price < 취소선 정가다. 같으면 봇 응답이라 할인이 붕괴된 것.
            if strike and price >= strike:
                return {"error": "즉시할인 붕괴(정가=취소선가)", "blocked": True, "title": title}
            return {"price": price, "vendor": "11번가", "method": "json-ld", "title": title}
        return {"error": "가격 패턴 미검출", "blocked": True, "title": title}

    if "danawa.com" in url:
        if re.search(r"가격비교\s*중지|판매\s*중지|생산\s*종료", html):
            return {"error": "가격비교중지/단종", "delisted": True, "title": title}
        m = re.search(r'class="prc_c">([\d,]{4,})', html)
        if not m:
            return {"error": "가격 패턴 미검출", "title": title}
        # 주의: fetch HTML에는 최저가 1건만 온다(몰 목록은 AJAX). 2위 비교는 여기서 불가.
        # 판매처는 텍스트(<span class="logo">디피컴</span>) 또는
        # 로고 이미지(<img alt='11번가'>) 둘 중 하나다. 둘 다 받아낸다.
        head = html[max(0, m.start() - 400):m.start()]
        v = re.search(r'<span class="logo">\s*([^<>]{1,40}?)\s*</span>', head) \
            or re.search(r"<span class=\"logo\">\s*<img[^>]*alt=['\"]([^'\"]{1,40})", head)
        d = re.search(r'id="topFixDeliveryFee"\s*>([^<]*)<', html)
        return {"price": int(m.group(1).replace(",", "")),
                "vendor": (v.group(1).strip() if v else None),
                "delivery": (d.group(1).strip() if d else None),
                "method": "danawa", "title": title}

    # 올서버: 상세페이지에 schema.org JSON-LD가 있어 fetch로 잡힌다.
    # 단 '가격문의' 상품은 파싱 실패가 아니라 **견적 품목이라는 사실**이다. 구분해서 기록한다
    # (이걸 '가격 패턴 미검출' 에러로 뭉개면 불투명 품목이 그냥 실패 로그로 사라진다)
    if "allserver.co.kr" in url:
        m = re.search(r'prdSalePrice"?\s*:\s*([\d]+)', html)
        price = int(m.group(1)) if m else 0
        if not price:
            m2 = re.search(r'"offers"\s*:\s*\{[^}]*"price"\s*:\s*([\d.]+)', html)
            price = int(float(m2.group(1))) if m2 else 0
        if price > 1000:
            return {"price": price, "vendor": "올서버", "method": "json-ld", "title": title}
        if "가격문의" in html:
            return {"error": "가격문의(견적품목)", "quote_required": True, "title": title,
                    "quote_source": "올서버 allserver.co.kr"}
        return {"error": "가격 패턴 미검출", "title": title}

    # 일반몰: JSON-LD → og:price → 흔한 클래스
    for m in re.finditer(r'"price"\s*:\s*"?([\d.,]+)"?', html):
        v = int(float(str(m.group(1)).replace(",", "")))
        if v > 1000:
            return {"price": v, "vendor": None, "method": "json-ld", "title": title}
    m = re.search(r'property=["\']product:price:amount["\']\s+content=["\']([\d.]+)', html)
    if m:
        return {"price": int(float(m.group(1))), "vendor": None,
                "method": "og:price", "title": title}
    m = re.search(r'class="[^"]*(?:prc_c|price_num|total_price|sale_price|productPrice)[^"]*"'
                  r'[^>]*>\s*([\d,]{5,})', html)
    if m:
        return {"price": int(m.group(1).replace(",", "")), "vendor": None,
                "method": "class", "title": title}
    # 국내 쇼핑몰 빌더(영문1·그늬·사이송)는 가격을 class가 아니라 id로 달고 나온다.
    m = re.search(r'id="(?:price|sit_sell_price|total_price|sit_tot_price)"[^>]*>\s*([\d,]{4,})', html)
    if m:
        return {"price": int(m.group(1).replace(",", "")), "vendor": None,
                "method": "id", "title": title}
    # 여기까지 오면 본문에 가격이 아예 없는 것(JS 렌더)이지 '링크가 죽은' 게 아니다.
    # 영구 실패로 묻지 말고 **브라우저 티어로 에스컬레이션**한다.
    # (2026-08-21: 온누리음향·윤사운드·뮤플이 여기 걸려 3일째 '실패'로만 남아 있었다.
    #  온누리음향은 브라우저로 열면 770,000₩이 멀짱하게 보인다)
    return {"error": "가격 패턴 미검출(JS 렌더 추정)", "blocked": True, "title": title}


# ── 라벨↔URL 불일치 검출 ─────────────────────────────────
CAP_PAT = re.compile(r"(\d+(?:\.\d+)?)\s*(TB|GB)\b", re.I)
BRANDS = ("삼성", "SK하이닉스", "하이닉스", "마이크론", "Micron", "WD", "Western Digital",
          "Seagate", "시게이트", "SK hynix", "Crucial", "ESSENCORE", "에센코어",
          "Kingston", "킹스턴", "ADATA", "Solidigm", "Intel", "AMD", "NVIDIA",
          "MSI", "ASUS", "GIGABYTE", "기가바이트", "ZOTAC", "PALIT", "HPE", "HP",
          "Dell", "Lenovo", "G.SKILL", "커세어", "CORSAIR")


def _caps(s):
    """문자열에서 용량 집합을 GB 단위로 뽑는다."""
    out = set()
    for v, u in CAP_PAT.findall(s or ""):
        gb = float(v) * (1024 if u.upper() == "TB" else 1)
        out.add(int(gb))
    return out


def _brands(s):
    low = (s or "").lower()
    return {b.lower() for b in BRANDS if b.lower() in low}


def label_mismatch(row_name, row_spec, title):
    """워크북 라벨과 실제 상품 페이지 제목이 다른 물건을 가리키는지 본다.

    2026-08-21에 임계초과 22건 중 다수가 시세 변동이 아니라 이것이었다:
    라벨 'DDR5-6400 ECC/REG 32GB'(245만)에 붙은 URL이 64GB(464.5만)였고,
    'SK hynix P41'에 삼성 9100 PRO URL이, 'SN850X 1TB'에 2TB URL이 붙어 있었다.
    가격이 아니라 **상품이 다른 것**이므로 시세 변동으로 보고하면 안 된다.
    """
    if not title:
        return None
    label = "%s %s" % (row_name or "", row_spec or "")
    lc, tc = _caps(label), _caps(title)
    if lc and tc and not (lc & tc):
        return "용량 불일치: 라벨 %s ↔ 페이지 %s" % (
            sorted(lc), sorted(tc))
    lb, tb = _brands(label), _brands(title)
    if lb and tb and not (lb & tb):
        return "브랜드 불일치: 라벨 %s ↔ 페이지 %s" % (sorted(lb), sorted(tb))
    return None


# ── 워크북 스캔 ───────────────────────────────────────────
def _parse_date(s):
    if not s:
        return None
    s = str(s)
    m = DATE_PAT.search(s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = MD_PAT.search(s)
    if m:
        try:
            return datetime.date(2026, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def scan(include_closed=False):
    """살아있는 공고의 물품 행을 모은다. 워크북은 절대 수정하지 않는다."""
    today = datetime.date.today()
    notices, rows = [], []
    for f in sorted(glob.glob(os.path.join(ROOT, "2026*/*.xlsx"))):
        if os.path.basename(f).startswith("~$"):
            continue
        try:
            wb = load_workbook(f)
        except Exception:
            continue
        ws = wb["분석"] if "분석" in wb.sheetnames else wb.active
        num = lambda v, d=0: v if isinstance(v, (int, float)) else d
        od = _parse_date(ws["B7"].value) or _parse_date(ws["F6"].value)
        live = (od is None) or (od >= today)
        if not live and not include_closed:
            continue
        base, rate, extra = num(ws["B14"].value), num(ws["B15"].value, 0.96), num(ws["B50"].value)
        n = {"file": os.path.basename(f), "title": ws["B4"].value, "bid_no": ws["B5"].value,
             "open_date": str(od or ""), "bid": round(base * rate) if base else 0,
             "extra": extra, "sel_cost": 0}
        for r in range(26, 46):
            nm, u = ws["A%d" % r].value, ws["E%d" % r].value
            if not nm:
                continue
            qty = num(ws["C%d" % r].value, 1) or 1
            price = num(ws["F%d" % r].value)
            sel = str(ws["G%d" % r].value or "").strip() == "O"
            if sel:
                n["sel_cost"] += qty * price
            k = product_key(str(u or ""))
            if k and u:
                rows.append({"file": n["file"], "row": r, "name": str(nm), "key": k,
                             "url": str(u).strip(), "qty": qty, "price": price, "sel": sel,
                             "spec": str(ws["B%d" % r].value or "")})
        notices.append(n)
    return notices, rows


# ── 메인 ──────────────────────────────────────────────────
def register_products(rows):
    """워크북에만 있고 products.json에 없던 상품을 등록한다.

    이전 구조의 구멍: 조사는 했는데 DB에 안 들어가 다음 공고에서 또 재조사하던 상품이
    40개였다. 이제 갱신기가 지나가면서 자동으로 주워 담는다.
    """
    products = load_products()
    added = 0
    for r in rows:
        if r["key"] in products:
            continue
        products[r["key"]] = {
            "name": r["name"], "brand": None,
            "model_code": extract_model_code(r.get("spec"), r["name"]),
            "spec": r.get("spec", ""), "qty_basis": "개당", "volatility": "normal",
            "urls": [r["url"]], "aliases": [],
            "note": "워크북에서 자동 등록 (price_refresh)",
            "first_seen": datetime.date.today().isoformat(), "source_file": r["file"],
        }
        added += 1
    if added:
        save_products(products)
    return added


def main():
    th = 0.5
    if "--threshold" in sys.argv:
        th = float(sys.argv[sys.argv.index("--threshold") + 1])
    include_closed = "--all" in sys.argv
    dry = "--dry" in sys.argv
    browser_in = None
    if "--browser-in" in sys.argv:
        browser_in = sys.argv[sys.argv.index("--browser-in") + 1]

    notices, rows = scan(include_closed)
    nmap = {n["file"]: n for n in notices}
    # 코어(2개 이상 공고 재등장) 표시용
    from collections import Counter
    freq = Counter(r["key"] for r in rows)
    targets = {}
    for r in rows:
        targets.setdefault(r["key"], r["url"])

    print("대상: 살아있는 공고 %d건 / 물품행 %d개 / 고유 상품 %d개%s"
          % (len(notices), len(rows), len(targets), "" if not include_closed else " (마감건 포함)"))

    now = datetime.datetime.now().isoformat(timespec="seconds")
    keys = list(targets)
    if browser_in:
        # 브라우저 티어 결과 합류 (봇차단 사이트). 역시 LLM 판단은 안 들어간다.
        incoming = json.load(open(browser_in, encoding="utf-8"))
        recs = [observe(r["key"], r.get("price"), vendor=r.get("vendor"), method="browser",
                        url=r.get("url") or targets.get(r["key"]),
                        page_title=r.get("title"),
                        listing_status=("delisted" if r.get("delisted") else "active"),
                        core=freq.get(r["key"], 0) >= 2)
                for r in incoming if r.get("price")]
        if not dry:
            append_obs(recs)
        result = {r["key"]: r for r in recs}
        el = 0.0
        print("브라우저 결과 %d건 합류 (fetch 생략)" % len(recs))
    else:
        t0 = time.time()
        with ThreadPoolExecutor(max_workers=4) as ex:
            got = list(ex.map(lambda k: (k, scrape(targets[k])), keys))
        result = dict(got)
        el = time.time() - t0
        ok = {k: v for k, v in result.items() if v.get("price")}
        print("갱신 완료: %.1f초 | 성공 %d/%d | 실패 %d  (LLM 호출 0)"
              % (el, len(ok), len(keys), len(keys) - len(ok)))
        if not dry:
            append_obs([observe(k, v.get("price"), vendor=v.get("vendor"),
                                method=v.get("method") or "fetch", url=targets[k],
                                core=freq[k] >= 2, delivery=v.get("delivery"),
                                error=v.get("error"), page_title=v.get("title"),
                                listing_status=("delisted" if v.get("delisted") else "active"))
                        for k, v in result.items()])

    # 영향 계산은 로그의 '상품키별 최신 관측'을 기준으로 한다
    # → 브라우저로 나중에 채운 값도 자동으로 반영된다.
    failures_now = {k: v for k, v in result.items() if not v.get("price")}
    merged = dict(latest_by_key())
    merged.update({k: v for k, v in result.items() if v.get("price")})
    result = merged

    # ── 예외 게이트: 마진율 영향으로 거른다 ──
    impact = {}
    for r in rows:
        v = result.get(r["key"])
        if not r["sel"] or not v or not v.get("price"):
            continue
        d = (v["price"] - r["price"]) * r["qty"]
        if d:
            # 워크북 단가가 0원인 행(미기입/견적대기)은 변동률이 정의되지 않는다.
            # 예전엔 여기서 ZeroDivisionError로 리포트 전체가 죽었다 (2026-08-25).
            base = r["price"]
            pct = ((v["price"] - base) / base * 100) if base else None
            impact.setdefault(r["file"], []).append({**r, "new": v["price"],
                                                     "vendor": v.get("vendor"), "delta": d,
                                                     "pct": pct})
    # ── 라벨↔URL 불일치를 먼저 걸러낸다 ──
    # 이걸 시세 변동으로 보고하면 매일 같은 가짜 급등을 다시 보게 된다.
    # 제목은 이번 fetch 결과 → 없으면 과거 관측에 기록된 page_title 순으로 본다.
    # (브라우저 합류 모드에서도 불일치 검출이 살아있게 하려고)
    titles = {}
    for k, v in result.items():
        t = v.get("title") or v.get("page_title")
        if t:
            titles[k] = t
    mismatches, mm_keys = [], set()
    for f, items in impact.items():
        for i in items:
            why = label_mismatch(i["name"], i.get("spec"), titles.get(i["key"]))
            if why:
                mm_keys.add((f, i["key"]))
                mismatches.append({"file": f, "name": i["name"], "url": i["url"],
                                   "key": i["key"], "old": i["price"], "new": i["new"],
                                   "pct": (round(i["pct"], 1) if i["pct"] is not None else None),
                                   "page_title": titles.get(i["key"]), "why": why})

    # 극단 변동은 시세가 아니라 URL이 다른 상품을 가리키는 경우가 많다.
    # 가격을 그대로 받아들이면 안 되고 상품 동일성부터 확인해야 하는 건이라 따로 뽑는다.
    SUSPECT_PCT = 40.0
    suspects = []
    for f, items in impact.items():
        for i in items:
            if (i["pct"] is None or abs(i["pct"]) >= SUSPECT_PCT) and (f, i["key"]) not in mm_keys:
                suspects.append({"file": f, "name": i["name"], "url": i["url"], "key": i["key"],
                                 "old": i["price"], "new": i["new"],
                                 "pct": (round(i["pct"], 1) if i["pct"] is not None else None),
                                 "vendor": i.get("vendor"), "page_title": titles.get(i["key"]),
                                 "why": "변동폭 과대 — URL이 다른 상품을 가리킬 가능성. 상품 동일성 먼저 확인"})

    # ── 미검증 가격은 마진에 넣지 않는다 (2026-08-25) ──
    # 라벨↔URL 불일치·상품동일성 의심 행은 "확인 전에는 쓰면 안 되는 값"이다.
    # 그런데 예전엔 그 값으로 마진을 다시 계산해 리포트 맨 위에 올렸다. 그 결과
    # 한밭대 -37%→-344%(같은 URL 7행 재사용), KISTI -1.8%→+48.4%(48GB 라벨에 24GB URL),
    # 계명대 11.6%→19.0%(판매종료된 다나와 최저가)처럼 **없는 판정 뒤집힘**이 매일 상단에 떴다.
    # 이제 margin_after는 의심 행을 워크북 값으로 되돌린 '보수 마진'이고,
    # 검증 전 참고용 수치는 margin_after_unverified에 따로 싣는다.
    unverified = mm_keys | {(x["file"], x["key"]) for x in suspects}
    flagged = []
    for f, items in impact.items():
        n = nmap[f]
        if not n["bid"]:
            continue
        d_all = sum(i["delta"] for i in items)
        d = sum(i["delta"] for i in items if (f, i["key"]) not in unverified)
        n_unv = sum(1 for i in items if (f, i["key"]) in unverified)
        c0 = n["sel_cost"] + n["extra"]
        m0 = (n["bid"] - c0) / n["bid"] * 100
        m1 = (n["bid"] - c0 - d) / n["bid"] * 100
        m1u = (n["bid"] - c0 - d_all) / n["bid"] * 100
        if abs(m1 - m0) >= th or n_unv:
            flagged.append({"file": f, "title": n["title"], "bid_no": n["bid_no"],
                            "open_date": n["open_date"], "bid": n["bid"],
                            "margin_before": round(m0, 1), "margin_after": round(m1, 1),
                            "delta_pp": round(m1 - m0, 1), "cost_delta": d,
                            "n_unverified": n_unv,
                            "margin_after_unverified": round(m1u, 1),
                            "items": items})
    fails = [{"key": k, "url": targets.get(k), **v} for k, v in failures_now.items()
             if k not in merged or not merged[k].get("price")]
    if browser_in:
        # 브라우저 모드는 fetch를 안 하므로 failures_now가 비어 있다.
        # 그대로 쓰면 직전 실행의 실패 목록이 통째로 지워진다(2026-08-21 실사고).
        # 이번에 해결된 키만 빼고 나머지는 보존한다.
        try:
            prev = json.load(open(EXC, encoding="utf-8")).get("failures", [])
        except Exception:
            prev = []
        # '해결'은 **이번 브라우저 입력에 가격이 들어온 키**만이다.
        # merged(과거 관측 포함)로 판정하면 예전에 한번 잡혔던 단종·판매종료 상품까지
        # "해결됨"으로 사라져 추적을 놓친다.
        solved = {r["key"] for r in incoming if r.get("price")}
        fails = [f for f in prev if f.get("key") not in solved]

    payload = {"ran_at": now, "threshold_pp": th, "n_products": len(keys),
               "n_ok": sum(1 for k in keys if result.get(k, {}).get("price")),
               "elapsed_sec": round(el, 1),
               "flagged": sorted(flagged, key=lambda x: x["delta_pp"]),
               "label_mismatch": mismatches,
               "suspect_identity": suspects,
               "failures": fails}
    if not dry:
        json.dump(payload, open(EXC, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        n_added = register_products(rows)
        print("\n상품 등록 +%d개 / 파생 뷰 %d행 재생성 (price-db.json)"
              % (n_added, len(rebuild_view())))

    print("\n■ AI 검토 대상 (마진율 %.1f%%p 이상 변동): %d건 / 전체 공고 %d건"
          % (th, len(flagged), len(notices)))
    for x in payload["flagged"]:
        note = ("   [미검증 {}행 제외 · 포함시 {:.1f}%]".format(x["n_unverified"], x["margin_after_unverified"])
                if x.get("n_unverified") else "")
        print("  {:44s}  {:.1f}% → {:.1f}% ({:+.1f}%p)  원가 {:+,}{}".format(
            x["file"][:44], x["margin_before"], x["margin_after"], x["delta_pp"], x["cost_delta"], note))
        for i in x["items"]:
            print("      · {:16s} {:>11,} → {:<11,} x{} ({}) {}".format(
                i["name"][:16], i["price"], i["new"], i["qty"],
                ("{:+.1f}%".format(i["pct"]) if i["pct"] is not None else "원가0원행"),
                i.get("vendor") or ""))
    if mismatches:
        print("\n■ ⚠ 라벨↔URL 불일치 %d건 — 시세 변동이 아니다. 워크북 URL을 고쳐야 한다" % len(mismatches))
        for s in mismatches:
            print("  {:>9}  {:>11,} → {:<11,}  {} | {}".format(
                ("{:+.1f}%".format(s["pct"]) if s["pct"] is not None else "0원행"),
                s["old"], s["new"], s["name"][:26], s["file"][:34]))
            print("      {}".format(s["why"]))
            print("      페이지: {}".format((s.get("page_title") or "")[:80]))

    if suspects:
        print("\n■ 상품 동일성 의심 %d건 (가격을 받아들이기 전 URL이 같은 상품인지 확인)" % len(suspects))
        for s in suspects:
            print("  {:>11}  {:>11,} → {:<11,}  {} | {}".format(
                ("{:+.1f}%".format(s["pct"]) if s["pct"] is not None else "0원행"),
                s["old"], s["new"], s["name"][:18], s["file"][:34]))
            print("      {}".format(s["url"][:100]))

    if fails:
        print("\n■ 스크레이핑 실패 %d건 (AI 확인 필요)" % len(fails))
        for f in fails:
            tag = "봇차단" if f.get("blocked") else ("단종/중지" if f.get("delisted") else "")
            print("  %-34s %s %s" % (f["key"][:34], f.get("error"), tag))
    print("\n로그: %s" % (OBS if not dry else "(dry-run, 미기록)"))


if __name__ == "__main__":
    main()
