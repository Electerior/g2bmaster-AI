# -*- coding: utf-8 -*-
"""2026-08-19 링크 보정 배치 — 검색/홈 URL을 실제 상품 상세 URL로 교체.

FIXES: (파일키, 행) -> dict(url, price, note_prefix)
 - url=None  → 실판매 상품 페이지 없음. URL 비우고 '견적필요'로 표기(가격은 참고가로 유지)
"""
import glob, os, sys
from openpyxl import load_workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gen_analysis import url_problem  # noqa

BASE = os.environ.get("BIDPIPE_ROOT") or os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LINK_FONT = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
D = lambda p: f"https://prod.danawa.com/info/?pcode={p}"
Q = "★견적필요(실판매 페이지 없음, 가격은 참고가) 견적처: "

FIXES = {
 "R26BK01684577": {  # 화학연 AI서버
   32: (None, None, Q+"아이티마야 itmaya.co.kr / 올서버 allserver.co.kr / 엑스디노드 — ASUS ESC4000A-E12 국내 리테일 미유통"),
 },
 "R26BK01684606": {  # 부산대 GPU서버
   30: (None, None, Q+"아이티마야·올서버 — DDR5-6400 ECC/REG 32GB 국내 리테일 미유통(다나와는 64GB만)"),
   32: (None, None, Q+"올서버·서버스토어 — Micron 7500 PRO 3.84TB U.2 국내 리테일 미유통"),
   33: (None, None, Q+"아이티마야·올서버 — Broadcom MegaRAID 9560-16i 국내 리테일 미유통"),
   34: (None, None, Q+"아이티마야·올서버 — I350-T4 'OCP3.0' 미유통. PCIe 카드형(인텔 I350-T4V2 88,600원)은 폼팩터 불일치"),
   35: (None, None, Q+"아이티마야·올서버 — Intel E810-XXVDA2 OCP3.0 국내 리테일 미유통"),
   36: (D("15688502"), 14750, "EFM ipTIME SFP-MMF10G 10G SFP+ MMF 광모듈. 가격 21,390→14,750 정정"),
 },
 "R26BK01685019": {32: (None, None, Q+"아이티마야·올서버·엑스디노드 — Xeon6 1S 2U 8베이 베어본 국내 리테일 미유통")},
 "R26BK01685088": {30: (None, None, Q+"아이티마야·올서버·엑스디노드 — Xeon6 1S 2U 8베이 베어본 국내 리테일 미유통")},
 "R26BK01685174": {  # 코리아텍 데스크톱
   26: ("https://www.apple.com/kr/shop/buy-mac/mac-studio/m3-ultra-%EC%B9%A9-28%EC%BD%94%EC%96%B4-cpu-60%EC%BD%94%EC%96%B4-gpu-96gb-%EB%A9%94%EB%AA%A8%EB%A6%AC-2tb-%EC%A0%80%EC%9E%A5-%EC%9E%A5%EC%B9%98",
        9840000, "Apple 공식 구성 URL(28C/60C·96GB·2TB) 확인 ₩9,840,000"),
   29: (D("11367381"), 279280, "삼성전자 포터블 SSD T7 (1TB)"),
   30: (D("69059687"), 507700, "인텔 코어 울트라7 시리즈2 265K"),
   31: (D("101955788"), 85250, "Thermalright FROZEN HORIZON 360 Digital ARGB 서린. 가격 163,900→85,250 정정(검색URL 기반 과대계상)"),
   32: (D("69696269"), 355930, "ASUS PRIME Z890-P-CSM 코잇"),
   33: (D("18911780"), 353400, "삼성전자 DDR5-5600 (PC5-44800)"),
   34: (D("86814989"), 750990, "MSI 지포스 RTX 5060 Ti 벤투스 2X OC 플러스 D7 8GB"),
   35: (D("18297002"), 402860, "삼성전자 990 PRO M.2 NVMe (1TB). 가격 358,900→402,860 정정"),
   36: (D("11884168"), 94580, "darkFlash DLX21 RGB MESH 강화유리"),
   37: (D("69468725"), 186000, "SuperFlower SF-1000F14GE LEADEX III GOLD UP ATX3.1"),
   38: (D("15449852"), 205500, "Microsoft Windows 11 Pro DSP(COEM) 64bit 한글. 가격 259,000→205,500 정정"),
 },
 "R26BK01685202": {27: (None, None, Q+"리더스시스템즈 1544-9880 / sales@leaderssys.com (NVIDIA 엘리트 파트너)")},
 "R26BK01685267": {28: (D("102618809"), 13299050, "ASUS Ascent GX10 (128GB/M.2 4TB) 대원씨티에스 유통, 옥션·G마켓 공식판매점")},
 "R26BK01686331": {28: (D("102618809"), 13299050, "ASUS Ascent GX10 (128GB/M.2 4TB) 대원씨티에스 유통, 옥션·G마켓 공식판매점")},
 "KETI_Q32": {26: ("https://www.comoasis.co.kr/shop/item.php?it_id=5749612822&service_id=pcdn", 26279980,
        "RTX PRO 6000 Blackwell 워크스테이션 에디션 D7 96GB (600W/2팬/304.8mm) — 컴오아시스. 가격 21,288,990→26,279,980 정정")},
 "KETI_Q33": {26: ("https://smartstore.naver.com/serverbiz/products/11962529128", 23045000,
        "RTX PRO 6000 Blackwell Max-Q 96GB (300W/블로워/266.7mm) — 서버비즈(에스티컴퓨터), OEM 정품·NVIDIA 3년보증. 가격 21,288,990→23,045,000 정정")},
 "R26BK01685910": {26: ("https://smartstore.naver.com/serverbiz/products/11962529128", 23045000,
        "RTX PRO 6000 Blackwell Max-Q 96GB (300W) — 서버비즈(에스티컴퓨터). B2B 02-707-0120. 가격 21,288,990→23,045,000 정정")},
 "LHK005133367": {26: (D("98183051"), 396490, "HP 오피스젯 프로 9730 와이드 포맷 (A3, 34ppm, 복사·스캔)")},
 "R26BK01685726": {  # 국보연 스마트기기
   26: (D("123621603"), 3244600, "삼성 갤럭시Z 폴드8 울트라 1TB 자급제(16GB/1TB). 가격 3,483,080→3,244,600 정정"),
   27: (D("123621618"), 2963400, "삼성 갤럭시Z 폴드8 1TB 자급제(16GB/1TB). 가격 3,152,600→2,963,400 정정"),
   29: (D("123236116"), 67540, "3종 합계: 신지모루 폴드8울트라 M-에어슬림 힌지(맥세이프) 19,150(pcode 123236116) + 폴드8 M-에어클로 베이직 12,500(123135172) + 아라리 플립8 에어로플렉스M 35,890(123670254)"),
 },
 "R26BK01685824": {  # 서울대 CryoEM
   27: (None, None, Q+"아이티마야·올서버 — AMD EPYC 9355 트레이/박스 국내 리테일 미유통(Dell CTO는 프리미엄 과대)"),
   29: (None, None, Q+"아이티마야·올서버 — 삼성 PM9A3 960GB U.2 국내 리테일 미유통"),
 },
 "R26BK01685839": {  # 대구대 PC 단가 15억
   30: (D("10174137"), 44830, "잘만 MegaMax 500W 80PLUS스탠다드. 가격 49,500→44,830 정정"),
   31: (D("28326653"), 33000, "에디 A1 MESH 강화유리 미들타워. 가격 36,700→33,000 정정"),
   35: (D("74835539"), 197000, "LG전자 24BA450 (FHD IPS 250nit, 피벗·스위블·엘리베이션, HDMI+DP+D-SUB) 포트조건 충족"),
   36: (D("122618297"), 373920, "알파스캔 콘퀘스트 32Q50G (QHD Fast IPS 400nit 1ms, 피벗·HAS, HDMI 2 + DP 1) 충족. 390,720→373,920 정정"),
   37: (D("108402173"), 299000, "크로스오버 34QW290GM (UWQHD VA 400nit 3000:1, HDMI 2 + DP 2, 엘리베이션) 충족. 369,000→299,000 정정 ※기존 후보 알파스캔 CU34G4는 300nit로 350cd 미달"),
 },
 "R26BK01685855": {31: (D("79408826"), None, "다나와 pcode 정상(오탐이었음) — cate 파라미터 제거")},
 "R26BK01686034": {31: (D("96483548"), 287890, "인터랙트 IW-S20 (무선 900MHz, 송+수신 2채널, XLR). 32ch PLL·True Diversity 표기는 상세페이지 미기재 → 확인필요")},
 "R26BK01686101": {27: (None, None, Q+"아이티마야·올서버 — NVIDIA A100 전용 8핀 전원 케이블은 서버 벤더 부속. 리테일 미유통")},
 "R26BK01686155": {28: (D("35234183"), 1890000, "스마트엑세스 충전보관함 L24-US3 (24구, PD3.0·PPS, 59kg). 가격 800,000→1,890,000 정정 ※잠금·이동식 여부 상세 미기재 → 확인필요")},
 "R26BK01686164": {  # 서울대 서버 1619
   27: (D("103130627"), 4644500, "삼성전자 DDR5-6400 ECC/REG (64GB). 가격 2,170,000→4,644,500 정정 ※서버 RDIMM 시황 급등 반영"),
   28: (D("17343065"), 383970, "삼성전자 PM893 벌크 (240GB) SATA 엔터프라이즈. 가격 300,000→383,970 정정"),
   29: (None, None, Q+"아이티마야·올서버·엑스디노드 — 4U 8-GPU 베어본(G494/ESC8000급) 국내 리테일 미유통"),
 },
 "R26BK01686771": {  # 가톨릭 PC
   32: (D("51815078"), 141540, "마이크로닉스 Classic II 850W 80PLUS골드 풀모듈러 ATX3.1. 샵다나와 리다이렉트 URL → 상품 URL로 교체, 149,000→141,540 정정"),
   33: (D("48648926"), 44540, "마이크로닉스 COOLMAX 스테디 미들타워. 샵다나와 리다이렉트 URL → 상품 URL로 교체, 49,900→44,540 정정"),
 },
}


def main():
    files = sorted(glob.glob(f"{BASE}/20260818/*.xlsx")) + sorted(glob.glob(f"{BASE}/20260819/*.xlsx"))
    done = 0
    for key, rowmap in FIXES.items():
        target = [f for f in files if key in os.path.basename(f)]
        if not target:
            print(f"!! 파일 못 찾음: {key}"); continue
        for path in target:
            wb = load_workbook(path); ws = wb["분석"]
            for r, (url, price, note) in rowmap.items():
                cell = ws.cell(row=r, column=5)
                old = cell.value
                if url:
                    cell.value = url; cell.hyperlink = url; cell.font = LINK_FONT
                else:
                    cell.value = ""; cell.hyperlink = None
                if price is not None:
                    ws.cell(row=r, column=6).value = price
                nc = ws.cell(row=r, column=9)
                prev = str(nc.value or "").strip()
                nc.value = (note + (" | " + prev if prev else ""))[:1000]
                done += 1
                print(f"  {os.path.basename(path)[:40]:42s} r{r} → {(url or '견적필요')[:58]}")
            wb.save(path)
    print(f"\n총 {done}행 보정 완료")


if __name__ == "__main__":
    main()
