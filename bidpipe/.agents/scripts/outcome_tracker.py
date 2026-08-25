# -*- coding: utf-8 -*-
"""개찰결과 회수 루프 (사후 대조).

분석 워크북(우리 산정값) ↔ 나라장터 개찰결과(실제 낙찰값)를 대조해
"우리 원가 추정이 실제와 몇 % 어긋났나"를 숫자로 만든다.

사용법
  python outcome_tracker.py pending            # 수집 대상(개찰 지난 건) 출력
  python outcome_tracker.py pending --all      # 미래 개찰 포함 전체 추적 목록
  python outcome_tracker.py merge results.json # 수집 결과 append (append-only)
  python outcome_tracker.py report             # 오차 리포트
"""
import sys, os, re, json, glob, datetime
from openpyxl import load_workbook

ROOT = os.environ.get("BIDPIPE_ROOT") or os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA = os.path.join(ROOT, ".agents/data/outcomes")
OUTCOMES = os.path.join(DATA, "outcomes.jsonl")
os.makedirs(DATA, exist_ok=True)

G2B_NO = re.compile(r"R\d{2}[A-Z]{2}\d{8}")
DATE_PAT = re.compile(r"(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})")
MD_PAT = re.compile(r"(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)")


def _parse_date(s, year_hint=2026):
    """워크북의 지저분한 일정 문자열에서 날짜만 뽑는다."""
    if not s:
        return None
    s = str(s)
    m = DATE_PAT.search(s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = MD_PAT.search(s)
    if m:
        try:
            return datetime.date(year_hint, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def _source_of(bid_no):
    """공고번호 형식으로 조달 시스템을 판정 (수집 경로가 다르다)."""
    if not bid_no:
        return "unknown"
    if G2B_NO.search(bid_no):
        return "g2b"
    if bid_no.startswith("Q"):
        return "keti"          # KETI 소액견적 (keti.bidsign.co.kr)
    if re.match(r"^(HCF|LHK|HRF|HFA)", bid_no):
        return "d2b"           # 국방전자조달
    return "other"


def scan_workbooks():
    """분석 워크북에서 우리가 산정한 값을 뽑는다."""
    out = []
    for f in sorted(glob.glob(os.path.join(ROOT, "2026*/*.xlsx"))):
        # 사용자가 Excel로 파일을 열어두면 ~$임시 락 파일이 생긴다(zip 아님) - 건너뛴 것
        if os.path.basename(f).startswith("~$"):
            continue
        wb = load_workbook(f, data_only=False)
        ws = wb["분석"] if "분석" in wb.sheetnames else wb.active
        g = lambda c: ws[c].value
        raw_no = str(g("B5") or "").strip()
        m = G2B_NO.search(raw_no)
        if m:
            bid_no = m.group(0)
        elif raw_no:
            # KETI Q..., 국방 HCF/LHK... : 앞 토큰에서 차수 접미사(-01)만 떼어낸다
            bid_no = re.sub(r"-\d{2,3}$", "", raw_no.split()[0])
        else:
            bid_no = None
        num = lambda v, d=0: v if isinstance(v, (int, float)) else d
        cost = 0
        for r in range(26, 46):
            if ws[f"A{r}"].value and str(ws[f"G{r}"].value or "").strip() == "O":
                # 사용자가 엑셀에서 직접 편집하면 수량/단가가 문자열('3대', '견적')로 들어올 수 있다
                cost += num(ws[f"C{r}"].value, 1) * num(ws[f"F{r}"].value, 0)
        base = num(g("B14"))
        rate = num(g("B15"), 0.96)
        extra = num(g("B50"))
        our_bid = round(base * rate) if base else None
        our_cost = cost + extra
        out.append({
            "file": os.path.relpath(f, ROOT),
            "bid_no": bid_no,
            "bid_no_raw": raw_no,
            "source": _source_of(bid_no or raw_no),
            "title": g("B4"),
            "org": g("F5"),
            "open_date": str(_parse_date(g("B7")) or _parse_date(g("F6")) or ""),
            "est_price": g("B13"),
            "base_price": base,
            "rate": rate,
            "our_bid": our_bid,
            "our_cost": our_cost,
            "our_margin": (our_bid - our_cost) if our_bid else None,
            "our_margin_rate": round((our_bid - our_cost) / our_bid * 100, 1) if our_bid else None,
        })
    return out


def load_outcomes():
    if not os.path.exists(OUTCOMES):
        return []
    rows = []
    with open(OUTCOMES, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def latest_by_no(rows):
    """append-only 로그에서 공고번호별 최신 관측을 뽑는다."""
    d = {}
    for r in rows:
        k = r.get("bid_no")
        if not k:
            continue
        if k not in d or (r.get("collected_at") or "") >= (d[k].get("collected_at") or ""):
            d[k] = r
    return d


def cmd_pending(show_all=False):
    today = datetime.date.today()
    wbs = scan_workbooks()
    done = latest_by_no(load_outcomes())
    rows = []
    for w in wbs:
        od = _parse_date(w["open_date"]) if w["open_date"] else None
        prev = done.get(w["bid_no"])
        # '직찰-영구미수집' = 입찰서제출·개찰이 외부 사이트(이비즈포유 등)에서 돌아
        # 나라장터 개찰결과에 영원히 안 올라온다. 재시도해도 소용없으니 대기열에서 뻐다.
        settled = prev and prev.get("status") in (
            "개찰완료", "유찰", "취소", "직찰-영구미수집")
        if settled:
            state = "수집완료"
        elif od and od > today:
            state = "개찰전"
        elif od:
            state = "★수집대상"
        else:
            state = "일정미상"
        if not show_all and state != "★수집대상":
            continue
        rows.append((state, w["bid_no"], w["source"], w["open_date"] or "-",
                     (w["title"] or "")[:34], w["our_bid"], w["our_cost"]))
    rows.sort(key=lambda r: (r[0], r[3]))
    print(f"{'상태':10s} {'공고번호':16s} {'경로':6s} {'개찰일':11s} {'공고명':36s} {'우리투찰':>12s} {'우리원가':>12s}")
    for s, no, src, od, t, b, c in rows:
        print(f"{s:10s} {str(no):16s} {src:6s} {od:11s} {t:36s} {str(b or '-'):>12s} {str(c or '-'):>12s}")
    print(f"\n총 {len(rows)}건" + ("" if show_all else " (수집대상만. 전체는 --all)"))


def cmd_merge(path):
    """수집 결과(JSON 배열)를 append-only 로그에 추가."""
    new = json.load(open(path, encoding="utf-8"))
    if isinstance(new, dict):
        new = [new]
    n = 0
    with open(OUTCOMES, "a", encoding="utf-8") as fh:
        for r in new:
            r.setdefault("collected_at", datetime.datetime.now().isoformat(timespec="seconds"))
            r.setdefault("source", "g2b-개찰결과")
            bidders = r.get("bidders") or []
            if bidders:
                win = min((b for b in bidders if b.get("amount")),
                          key=lambda b: b.get("rank") or 999, default=None)
                r.setdefault("win_amount", win.get("amount") if win else None)
                r.setdefault("win_company", win.get("company") if win else None)
                r.setdefault("n_bidders", len(bidders))
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")
            n += 1
    print(f"{n}건 append → {OUTCOMES}")


def cmd_report():
    wbs = {w["bid_no"]: w for w in scan_workbooks() if w["bid_no"]}
    outs = latest_by_no(load_outcomes())
    joined = []
    for no, o in outs.items():
        w = wbs.get(no)
        if not w:
            continue
        win = o.get("win_amount")
        row = {
            "bid_no": no, "title": (w["title"] or "")[:30], "status": o.get("status"),
            "n_bidders": o.get("n_bidders"), "our_cost": w["our_cost"],
            "our_bid": w["our_bid"], "win": win,
            "g2b_est": o.get("est_price"), "our_base": w["base_price"],
        }
        if win and w["our_cost"]:
            row["cost_vs_win"] = round(w["our_cost"] / win * 100, 1)   # 낙찰가 대비 우리 원가 %
            row["margin_if_won"] = win - w["our_cost"]
        if win and w["our_bid"]:
            row["would_win"] = w["our_bid"] < win
        if o.get("est_price") and w["base_price"]:
            row["base_err"] = round((w["base_price"] - o["est_price"]) / o["est_price"] * 100, 1)
        joined.append(row)

    if not joined:
        print("대조 가능한 건 없음. 개찰 후 수집하면 채워진다.")
        print(f"  분석 워크북 {len(wbs)}건 / 수집된 개찰결과 {len(outs)}건")
        return

    print(f"{'공고번호':16s} {'상태':8s} {'참여':>4s} {'우리원가':>12s} {'낙찰가':>12s} "
          f"{'원가/낙찰가':>10s} {'낙찰시마진':>12s} {'공고명'}")
    for r in sorted(joined, key=lambda x: x.get("cost_vs_win") or 999):
        print(f"{r['bid_no']:16s} {str(r['status'] or '-'):8s} {str(r['n_bidders'] or '-'):>4s} "
              f"{r['our_cost']:>12,} {str(r.get('win') or '-'):>12} "
              f"{str(r.get('cost_vs_win') or '-'):>10} {str(r.get('margin_if_won') or '-'):>12} {r['title']}")

    ratios = [r["cost_vs_win"] for r in joined if r.get("cost_vs_win")]
    if ratios:
        ratios.sort()
        med = ratios[len(ratios) // 2]
        print(f"\n■ 원가추정 신뢰도 지표 (n={len(ratios)})")
        print(f"  낙찰가 대비 우리 원가 중앙값: {med}%   (100% 초과 = 그 가격엔 못 맞춘다는 뜻)")
        print(f"  최저 {min(ratios)}% / 최고 {max(ratios)}%")
        wins = [r for r in joined if r.get("would_win") is True]
        print(f"  우리 투찰가가 실제 낙찰가보다 낮았을 건: {len(wins)}/{len([r for r in joined if r.get('would_win') is not None])}")
    errs = [r["base_err"] for r in joined if r.get("base_err") is not None]
    if errs:
        errs.sort()
        print(f"  기초금액 대비 실제 예정가격 오차 중앙값: {errs[len(errs)//2]}%")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "report"
    if cmd == "pending":
        cmd_pending("--all" in sys.argv)
    elif cmd == "merge":
        cmd_merge(sys.argv[2])
    else:
        cmd_report()
