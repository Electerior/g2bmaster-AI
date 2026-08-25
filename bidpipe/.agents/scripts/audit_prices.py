# -*- coding: utf-8 -*-
"""분석 워크북의 '가격 근거'를 감사한다 (AGENTS.md 규칙 1·6 강제용).

마진율만 보면 안 되는 이유: 원가의 절반이 URL 없는 추정가면 그건 마진율이 아니다.
이 스크립트는 각 파일의 **추정가 비중**을 계산해 마진율 옆에 붙여주고,
보고 금지(REPORT-BLOCK) 대상을 골라낸다.

    python .agents/scripts/audit_prices.py '20260820/*.xlsx'
    python .agents/scripts/audit_prices.py '20260820/*.xlsx' --json
    python .agents/scripts/audit_prices.py '20260820/*.xlsx' --rows   # 문제 행까지

종료코드: 보고 불가 파일이 하나라도 있으면 1.
"""
import sys, glob, json, os
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gen_analysis import url_problem, upper_bound_problem

R0, R1 = 26, 45
EST_SHARE_BLOCK = 30.0          # 원가의 30% 이상이 추정이면 마진율 보고 금지


def _layout(ws):
    """물품행 끝행·부대비용 행을 동적으로 찾는다.
    2026-08-21 gen_analysis가 20행 초과 공고(충북대 서버 82행)에서 행을 늘리게 되면서
    26~45 / B50 고정 가정이 깨졌다(부대비용 자리에 문자열이 있어 TypeError)."""
    end = R1
    for r in range(R0, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == "상품가격 합계":
            end = r - 1
            break
    extra_row = None
    for r in range(end + 1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().startswith("부대비용"):
            extra_row = r
            break
    return end, extra_row


def audit_file(f):
    wb = load_workbook(f)
    if "분석" not in wb.sheetnames:
        return None
    ws = wb["분석"]
    base = ws["B14"].value or 0
    rate = ws["B15"].value
    if rate is None:
        rate = 0.96
    R1_, extra_row = _layout(ws)
    extra = (ws.cell(row=extra_row, column=2).value if extra_row else 0) or 0
    if not isinstance(extra, (int, float)):
        extra = 0

    sel_sum = est_amt = zero_quote = 0
    bad = []
    for r in range(R0, R1_ + 1):
        name = ws.cell(row=r, column=1).value
        if not name or ws.cell(row=r, column=7).value != "O":
            continue
        # 수량·단가에 문자열이 들어있으면(엑셀 수기 편집으로 공백 한 칸이 남는 일이 있다)
        # `4 * ' '`가 문자열 반복이 돼 `sel_sum += line`에서 TypeError로 죽는다.
        # 그랬던 이유로 2026-08-21까지 `audit_prices.py '날짜폴더/*.xlsx'`가
        # 네 폴더 전부 크래시했다 — 규칙 6이 보고 근거로 삼는 그 명령이다.
        # 숫자가 아니면 0으로 내리고 문제 행으로 드러낸다(조용히 넘기지 않는다).
        def _num(v, dflt=0):
            return v if isinstance(v, (int, float)) and not isinstance(v, bool) else dflt
        raw_qty = ws.cell(row=r, column=3).value
        raw_price = ws.cell(row=r, column=6).value
        qty = _num(raw_qty, 1) if raw_qty not in (None, "") else 1
        price = _num(raw_price, 0)
        if (raw_qty not in (None, "") and _num(raw_qty, None) is None) or \
           (raw_price not in (None, "") and _num(raw_price, None) is None):
            bad.append(dict(row=r, name=str(name)[:46], amount=0, kind="숫자아님",
                            reason=f"수량={raw_qty!r} 단가={raw_price!r} — 셀 값이 숫자가 아니다"))
        line = qty * price
        sel_sum += line
        url = (ws.cell(row=r, column=5).value or "").strip()
        vendor = str(ws.cell(row=r, column=4).value or "")
        note = str(ws.cell(row=r, column=9).value or "")
        # 상한값(조달 계약단가 등)을 원가 행에 쓴 경우. URL은 멀정해도 원가가 아니다
        ub = upper_bound_problem(url, vendor, note)
        if ub and line:
            est_amt += line
            bad.append(dict(row=r, name=str(name)[:46], amount=line,
                            kind="상한가사용", reason=ub))
            continue
        # (2026-08-21) 앟커 URL을 달고 있는 견적 행을 '전건 실가'로 오판하던 구멍을 막는다.
        # ETRI EA20261919·KISTI SSD가 실제로 이 경로로 감사를 통과했다.
        # (2026-08-21 2차) 구멍이 두 개 더 있었다 — 송원대 R26BK01686155가 이 경로로
        # '전건 실가 / 마진율 96.4% / 보고가능'을 받았다(선택 5행 중 3행이 0원 견적).
        #   ① 마커가 **품명 칸**에만 있는 행(`[견적필요] …`)을 아예 안 봤다
        #   ② 비고는 "견적 필요"(공백 O)만 찾는데 워크북 표기는 "견적필요"(공백 X)다
        # 공백을 지운 뒤 품명·판매처·비고를 모두 본다. 단 bare "견적"으로 넓힐 수는 없다
        # — 비고에 '견적 별도' 같은 문구를 단 실가 행까지 오탐하므로 마커 토큰만 쓴다.
        _blob = "".join(str(name).split()) + "".join(vendor.split()) + "".join(note.split())
        quoted = ("견적" in vendor) or any(t in _blob for t in ("견적필요", "견적처"))
        # 견적 미확보를 0원으로 넣은 행은 어느 게이트에도 안 잡힌다(line=0).
        # 그러면 총원가는 '하한'이고 마진율도 난관치다 — 별도로 세서 보고에 노출한다.
        if quoted and not line:
            zero_quote += 1
            continue
        if quoted and line:
            est_amt += line
            bad.append(dict(row=r, name=str(name)[:46], amount=line, kind="견적필요",
                            reason="vendor/비고에 견적 표기 — 기재 금액은 앟커일 수 있음"))
            continue
        # 사내 매입가표(티어 A)는 URL이 없어도 확정 원가다 (price-research "URL 규칙" 예외)
        if any(k in vendor for k in ("당사 매입", "당사매입", "매입가", "사내")):
            continue
        prob = url_problem(url)
        if prob and line:
            est_amt += line
            bad.append(dict(row=r, name=str(name)[:46], amount=line,
                            kind="추정가우회", reason=prob))

    bid = round(base * rate)
    # 규칙 15 (2026-08-21 사용자 확정): 부대비용은 원가에서 제외한다.
    #   부대비용은 정의상 추정치라 규칙 1(추정가 금지)과 충돌하고, 건별로 자의적이라
    #   공고 간 마진율 비교를 깨뜨린다. 마진율 기준은 '상품가격 합계'만이다.
    #   gen_analysis.py는 이 규칙을 지켰지만 이 감사기가 안 지켜서, 규칙이 생긴 뒤에도
    #   기존 워크북 51건(총 9,840만원)의 부대비용이 원가에 그대로 남아 대시보드·시트·보고에
    #   흘러들고 있었다. 세종대·서울과기대는 이것 때문에 흑자가 적자로 뒤집혀 보였다.
    #   (2026-08-21 시트 마이그레이션 중 물품합 vs 공고원가 교차검증으로 발견)
    #   값은 버리지 않고 extra로 실어 보낸다 — 원가가 아니라 리스크 항목으로 다룬다.
    cost = sel_sum
    if not bid:
        return None
    share = round(est_amt / cost * 100, 1) if cost else 0.0
    bypass = [b for b in bad if b["kind"] in ("추정가우회", "상한가사용")]
    return dict(
        file=os.path.basename(f), bid=bid, cost=cost, margin=bid - cost,
        margin_rate=round((bid - cost) / bid * 100, 1),
        extra_cost=extra,                # 원가 아님. 규칙 15 — 개요의 리스크 항목으로 다룬다
        est_rows=len(bad), est_amount=est_amt, est_share=share,
        zero_quote_rows=zero_quote,      # 원가 0원으로 들어간 견적 행 = 총원가가 하한이라는 뜻
        bypass_rows=len(bypass),
        # 견적 미확보 행을 0원으로 넣으면 총원가는 '하한'이고 마진율은 '상한'이다.
        # 그 숫자는 마진율이 아니므로 보고 대상에서 미리 밀어낸다 (규칙 6 판정보류).
        reportable=(share < EST_SHARE_BLOCK and not bypass and not zero_quote),
        rows=bad,
    )


def main(pattern, as_json=False, show_rows=False):
    # 사용자가 Excel로 파일을 열어두면 ~$임시 락 파일이 생긴다(zip 아님).
    # 스캔기(scan_dashboard/price_refresh/outcome_tracker)와 동일하게 건너뛴다.
    # (2026-08-24: 이 체크가 없어 KISTI 파일을 열어둔 채 '20260824/*.xlsx'를
    #  돌리면 BadZipFile로 네 폴더 전체가 크래시했다 — 규칙 6의 보고 근거 명령이
    #  파일 하나를 여는 순간 죽는다는 뜻이다.)
    files = [f for f in sorted(glob.glob(pattern))
             if not os.path.basename(f).startswith("~$")]
    res = [a for a in (audit_file(f) for f in files) if a]
    res.sort(key=lambda x: -x["margin_rate"])
    if as_json:
        print(json.dumps(res, ensure_ascii=False, indent=1))
    else:
        # '전 품목 실가'는 추정행이 없는 것만으로는 부족하다 — 0원 견적행은
        # 가격을 '추정'한 게 아니라 아예 안 잡힌 것이라 더 나쁜 상태다.
        ok = sum(1 for r in res if r["est_rows"] == 0 and not r["zero_quote_rows"])
        print(f"{len(res)}건 중 전 품목 실가 확보 {ok}건 / 보고가능 {sum(1 for r in res if r['reportable'])}건\n")
        print(f'{"공고":34s} {"투찰가":>13s} {"원가":>13s} {"마진율":>8s} {"가격근거":>22s}')
        for r in res:
            n = r["file"].replace(".xlsx", "")
            n = n.split("_", 1)[1] if "_" in n else n
            if r["est_rows"] == 0:
                basis = "전건 실가" if not r["zero_quote_rows"] else "원가미확정"
            else:
                basis = f'추정 {r["est_rows"]}행 {r["est_share"]}%'
                if r["bypass_rows"]:
                    basis += f' (우회 {r["bypass_rows"]})'
            if r["zero_quote_rows"]:
                basis += f' +견적0원 {r["zero_quote_rows"]}행'
            mark = "" if r["reportable"] else "  ← 보고금지"
            mr = f'{r["margin_rate"]}%' + ("*" if r["est_rows"] else
                                           ("↓" if r["zero_quote_rows"] else ""))
            print(f'{n[:34]:34s} {r["bid"]:>13,} {r["cost"]:>13,} {mr:>8s} {basis:>22s}{mark}')
            if show_rows:
                for b in r["rows"]:
                    print(f'      r{b["row"]:<3d} [{b["kind"]}] {b["name"]} {b["amount"]:,}원 — {b["reason"]}')
        print("\n* = 추정가 포함된 잠정치. 보고 시 마진율 단독 표기 금지(규칙 6).")
        print("↓ = 견적 미확보 행을 0원으로 넣은 건. 총원가는 하한이고 실제 마진은 표시보다 나쁘다.")
    return 1 if any(not r["reportable"] for r in res) else 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    sys.exit(main(args[0], "--json" in sys.argv, "--rows" in sys.argv))
