# -*- coding: utf-8 -*-
"""Format.xlsx 기반 공고 분석 파일 생성기"""
import shutil, sys, json, os, re, datetime
from urllib.parse import urlparse, parse_qs
from openpyxl import load_workbook
from openpyxl.styles import Font

_BIDPIPE_ROOT = os.environ.get("BIDPIPE_ROOT") or os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
FORMAT = os.path.join(_BIDPIPE_ROOT, "Format.xlsx")
LINK_FONT = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")

# ── URL 품질 검증 (2026-08-19 사용자 지적: 다나와 검색/홈 URL이 걸려서 상품에 못 감) ──
# 상품 상세 URL만 허용. 검색결과·카테고리·홈은 ERROR.
_SEARCH_PAT = re.compile(
    r"(dsearch\.php|/dsearch|/search[/?]|/search$|searchKeyword=|[?&]kwd=|[?&]query="
    r"|[?&]keyword=|product-category|/category/|[?&]cate=|[?&]category|_list\.php|/list[/?]"
    r"|/goods/list|search\.shopping|/brand/|/plusstore/)", re.I)
_ID_PAT = re.compile(r"\d{3,}")            # 상품 상세 URL은 대개 3자리 이상 상품ID를 가짐
_SLUG_PAT = re.compile(r"[a-z0-9]+(?:-[a-z0-9]+){2,}", re.I)  # 또는 하이픈 슬러그
_CODE_PAT = re.compile(r"/(?:dp|gp/product)/[A-Z0-9]{10}|/[A-Za-z0-9]*\d[A-Za-z0-9]{7,}")  # ASIN 등 영숫자 상품코드

# 제조사 직판 스토어는 상품ID 대신 제품라인명 경로를 쓴다(예: apple.com/kr/shop/buy-mac/mac-studio).
# 그 페이지에서 구성별 정가가 그대로 검증되므로 '상품 상세 URL' 요건을 충족한다.
# 오탐을 남겨두면 게이트 자체를 무시하게 되어 더 위험하다. (2026-08-20)
_VENDOR_STORE = re.compile(
    r"(apple\.com/[a-z-]+/shop/buy-|lge\.co\.kr/(?:care-accessories|[a-z-]+)/"
    r"|samsung\.com/sec/[a-z-]+/\S+/\S+|dell\.com/\S+/shop/\S+/spd/|hp\.com/\S+/shop/pdp/)", re.I)


# ── 상한값(bound=upper) 오용 차단 (2026-08-20) ────────────────────
# 조달 종합쇼핑몰 계약단가는 '우리가 이겨야 하는 천장'이지 우리 원가가 아니다.
# 그걸 원가 행에 넣으면 URL이 멀정해도 결과는 추정가와 같다 (마진이 과소평가된다).
_UPPER_HOSTS = ("shop.g2b.go.kr", "shopping.g2b.go.kr")


def upper_bound_problem(url, vendor=None, note=None):
    """상한값을 원가로 쓴 흔적이면 사유 문자열, 아니면 None"""
    u = str(url or "")
    if any(h in u for h in _UPPER_HOSTS):
        return ("조달 종합쇼핑몰 계약단가는 상한(bound=upper)다 — 원가 행에 쓸 수 없음. "
                "실판매가/견적가로 대체하고 이 값은 천장·예산근거로만 참조")
    v = str(vendor or "")
    if "계약단가" in v or "상한" in v:
        return "vendor가 상한값 출처(%s) — 원가 근거로 사용 불가" % v[:30]
    return None


def _is_internal_purchase(vendor):
    """사내 매입가표(티어 A) 행인지. URL 면제 대상.

    audit_prices.py와 동일한 토큰을 쓴다 — 한쪽만 고치면 게이트와 감사가 갈라진다.
    """
    v = str(vendor or "")
    return any(k in v for k in ("당사 매입", "당사매입", "매입가", "사내"))


def url_problem(u):
    """상품 상세 URL이 아니면 사유 문자열, 정상이면 None"""
    u = (u or "").strip()
    if not u:
        return "URL 없음"
    if not u.lower().startswith("http"):
        return "http(s) URL 아님"
    p = urlparse(u)
    q = parse_qs(p.query)
    path = p.path or ""
    # 다나와는 pcode 유무로 먼저 판정 (?pcode=123&cate=456 같은 정상 URL 오탐 방지)
    if "danawa.com" in p.netloc:
        code = (q.get("pcode") or q.get("code") or q.get("prodCode") or [""])[0]
        if code.strip():
            return None
        if _SEARCH_PAT.search(u) or path.strip("/") in ("", "info", "main"):
            return "다나와 검색/목록 URL — prod.danawa.com/info/?pcode=XXXX 형태 필요"
        return "다나와 URL에 pcode 없음 — prod.danawa.com/info/?pcode=XXXX 형태 필요"
    if _VENDOR_STORE.search(u):        # 제조사 공식몰 제품 페이지는 상품ID 없어도 인정
        return None
    if _SEARCH_PAT.search(u):
        return "검색결과/카테고리 URL — 상품 상세 URL 필요"
    if path.strip("/") == "" and not p.query:
        return "사이트 홈 URL — 상품 상세 URL 필요"
    if not (_ID_PAT.search(path + "?" + p.query) or _SLUG_PAT.search(path)
            or _CODE_PAT.search(path)):
        return "상품 식별자(ID/슬러그) 없는 랜딩 URL로 보임 — 상품 상세 URL 확인 필요"
    return None
# 2026-08-19 구조 변경: 분석/ → 날짜 폴더(YYYYMMDD)별 저장
OUTDIR = os.environ.get("GEN_OUTDIR") or os.path.join(
    _BIDPIPE_ROOT, "out", datetime.date.today().strftime("%Y%m%d"))
os.makedirs(OUTDIR, exist_ok=True)

ITEM_START, ITEM_END = 26, 45      # Format.xlsx의 물품 내역 행 범위


def _expand_item_rows(ws, n_items):
    """물품 행이 20행을 넘으면 45행 아래로 행을 늘리고 하단 수식·병합범위를 재작성한다.
    openpyxl의 insert_rows는 수식도 병합범위도 자동 보정하지 않으므로 직접 처리한다.
    반환값: 늘어난 행 수(extra)."""
    extra = max(0, n_items - (ITEM_END - ITEM_START + 1))
    if not extra:
        return 0
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange
    moved = [str(m) for m in ws.merged_cells.ranges if m.min_row > ITEM_END]
    for ref in moved:
        ws.unmerge_cells(ref)
    ws.insert_rows(ITEM_END + 1, extra)
    for ref in moved:
        cr = CellRange(ref); cr.shift(row_shift=extra)
        ws.merge_cells(str(cr))
    # 새로 생긴 물품 행에 선택금액 수식 부여 (원본 26~45행과 동일)
    for rr in range(ITEM_END + 1, ITEM_END + 1 + extra):
        ws[f"H{rr}"] = f'=IF($G{rr}="O",$C{rr}*$F{rr},0)'
    last = ITEM_END + extra
    s = 46 + extra                      # 상품가격 합계 행
    ws[f"H{s}"] = f"=SUM(H{ITEM_START}:H{last})"
    ws[f"B{49 + extra}"] = f"=H{s}"                                     # 상품가격 합
    ws[f"B{51 + extra}"] = f"=B{49 + extra}+B{50 + extra}"              # 총 원가
    ws[f"B{52 + extra}"] = "=B16"                                       # 투찰가
    ws[f"B{53 + extra}"] = f"=B{52 + extra}-B{51 + extra}"              # 마진
    ws[f"B{54 + extra}"] = f"=IF(B{52 + extra}=0,0,B{53 + extra}/B{52 + extra})"
    ws[f"B{56 + extra}"] = f'=IF(B{54 + extra}>=B{55 + extra},"검토 대상","제외")'
    return extra

def generate(meta, items, out_name):
    """meta: 공고정보 dict, items: [{name, spec, qty, vendor, url, price, sel, note, compliant}], out_name: 파일명
    compliant: RFP 행 대조 결과 '충족'/'확인필요'/'X'. sel=O 행은 반드시 기재 (미기재 시 WARN, X인데 O이면 ERROR)."""
    out = f"{OUTDIR}/{out_name}"
    shutil.copy(FORMAT, out)
    wb = load_workbook(out)
    ws = wb["분석"]
    # 1. 공고 정보
    ws["B4"] = meta["title"]
    ws["B5"] = meta["id"];          ws["F5"] = meta["org"]
    ws["B6"] = meta["date"];        ws["F6"] = meta["deadline"]
    ws["B7"] = meta.get("open", "");ws["F7"] = meta.get("delivery", "")
    ws["B8"] = meta["contract"];    ws["F8"] = meta["award"]
    ws["B9"] = meta.get("quali", "")
    ws["B10"] = meta.get("place", "");  ws["F10"] = meta.get("url", "")
    # 2. 금액
    ws["B13"] = meta["est_price"]   # 추정가격
    ws["B14"] = meta["base_price"]  # 기초금액(배정예산)
    ws["B15"] = meta.get("rate", 0.96)
    # 3. 개요
    ws["A19"] = meta["summary"]
    # 4. 물품 내역 (26행부터)
    # 2026-08-21: Format의 물품행은 26~45(20행)뿐이라 다품목 공고(충북대 서버 11품목=68행)에서
    #   행이 조용히 잘렸다. 잘린 행 = 원가에서 사라진 부품이므로 규칙 2(묶음 행 금지)의 취지를 정면으로 깬다.
    #   items가 20행을 넘으면 45행 아래로 행을 삽입하고 하단 수식·병합범위를 전부 재작성한다.
    extra = _expand_item_rows(ws, len(items))
    r = 26
    for it in items:
        if r > 45 + extra: break
        ws[f"A{r}"] = it["name"]
        ws[f"B{r}"] = it.get("spec", "")
        ws[f"C{r}"] = it.get("qty", 1)
        ws[f"D{r}"] = it.get("vendor", "")
        # URL은 반드시 '클릭되는' 하이퍼링크로 기록한다 (텍스트만 넣으면 Excel이 링크로 안 만듦)
        u = (it.get("url", "") or "").strip()
        cell = ws[f"E{r}"]
        cell.value = u
        if u.lower().startswith("http"):
            cell.hyperlink = u
            cell.font = LINK_FONT
        ws[f"F{r}"] = it.get("price", 0)
        ws[f"G{r}"] = "O" if it.get("sel") else "X"
        ws[f"I{r}"] = it.get("note", "")
        r += 1
    # 5. 부대비용 — 2026-08-21 사용자 지시로 **원가에서 제외**한다.
    #  이유: 부대비용은 정의상 추정치라 규칙 1(추정가 금지)과 충돌하고,
    #  건별로 자의적이라 공고 간 마진율 비교가 깨진다. 이행비용은 원가가 아니라
    #  계약조건 리스크(규칙 11)로 다룬다. 견적이 붙는 실물 항목(보증연장·설치지원 등)은
    #  부대비용이 아니라 **물품 행**으로 넣는다.
    _ec = meta.get("extra_cost", 0) or 0
    if _ec:
        print(f"NOTE:  {out_name} 부대비용 {_ec:,}원 입력됨 → 0으로 강제(2026-08-21 규칙). "
              f"근거는 개요의 리스크 항목으로 남길 것", file=sys.stderr)
    ws[f"B{50 + extra}"] = 0
    wb.save(out)
    # 6~7. 생성 게이트 (2026-08-20 개편)
    #  이전엔 ERROR를 stderr로 찍고 stdout으로는 깨끗한 margin_rate만 돌려줘서,
    #  보고에 쓰이는 stdout에는 경고가 안 남았다 → URL 없는 "참고 산정가"가 그대로
    #  마진율로 보고된 사고(30건 중 28건). 이제 품질 지표를 반환값에 같이 싣는다.
    blocking, provisional = [], []
    est_rows = est_amount = zero_quote = 0
    for it in items:
        if not it.get("sel"):
            continue
        line = it.get("price", 0) * it.get("qty", 1)
        if it.get("quote"):                       # 실판매가가 존재하지 않는 정당한 견적 항목
            if line:
                est_rows += 1
                est_amount += line
                provisional.append(f"[{it['name']}] 견적미확보 금액 {line:,}원이 원가에 포함됨")
            else:
                # (2026-08-21) 게이트의 큰 구멍: 견적을 **0원으로** 넣으면 est_rows가 안 올라
                # `all_verified:true` + `gate:OK` + 깨끗한 margin_rate가 그대로 나갔다.
                # 송원대 R26BK01686155가 이 경로로 선택 6행 중 3행이 견적 미확보인데
                # '마진율 96.4% / 전건 검증 완료'를 받았다.
                # 0원 견적행이 있으면 총원가는 '하한', 마진율은 '상한'이다 — 마진율이 아니다.
                zero_quote += 1
                provisional.append(f"[{it['name']}] 견적 미확보를 0원으로 계상 — 총원가는 하한이고 "
                                   f"마진율은 상한이다(보고불가)")
            if not str(it.get("quote_source", "") or "").strip():
                blocking.append(f"[{it['name']}] quote=true인데 quote_source(견적처) 미기재")
            # 견적은 만료된다. 유효기간이 없거나 지난 견적은 근거가 아니라 기억이다
            vu = str(it.get("quote_valid_until") or "").strip()[:10]
            if vu and vu < datetime.date.today().isoformat():
                blocking.append(f"[{it['name']}] 견적 유효기간 만료({vu}) — 재견적 전 원가로 쓸 수 없음")
            elif not vu:
                provisional.append(f"[{it['name']}] 견적 유효기간(quote_valid_until) 미기재")
        elif _is_internal_purchase(it.get("vendor")):
            # 사내 매입가표(티어 A)는 URL이 없어도 확정 원가다.
            # (price-research "URL 규칙" 예외 / audit_prices._layout와 같은 판정)
            # 2026-08-25: audit_prices에만 이 예외가 있어 두 도구가 어긋나 있었다.
            #   당사 GPU 매입가(URL 없음)를 넣으면 gen은 BLOCKED, audit은 '전건 실가'로
            #   서로 다른 답을 냈다. 같은 규칙을 쓰도록 통일한다.
            pass
        elif not it.get("url_ok"):
            prob = url_problem(it.get("url"))
            if prob:                              # ← 이게 "참고 산정가" 우회 경로다. 차단.
                est_rows += 1
                est_amount += line
                blocking.append(f"[{it['name']}] URL 불량 — {prob} (값: {it.get('url','')!r})")
        ub = upper_bound_problem(it.get("url"), it.get("vendor"), it.get("note"))
        if ub:
            blocking.append(f"[{it['name']}] {ub}")
        comp = str(it.get("compliant", "") or "").strip()
        if comp in ("X", "미달"):
            blocking.append(f"[{it['name']}] 규격미달(compliant=X)인데 sel=O — sel=X로 바꾸고 원가에서 제외")
        elif not comp:
            provisional.append(f"[{it['name']}] compliant 미기재 — RFP 전 행(특히 I/O PORTS) 대조 필요")

    # 검증용 마진 계산 (openpyxl은 수식 미계산이므로 파이썬으로 병행 계산)
    sel_sum = sum(it.get("price",0) * it.get("qty",1) for it in items if it.get("sel"))
    bid = round(meta["base_price"] * meta.get("rate", 0.96))
    cost = sel_sum                      # 부대비용 제외 (2026-08-21)
    margin = bid - cost
    mrate = margin / bid if bid else 0
    est_share = round(est_amount / cost * 100, 1) if cost else 0.0

    for m in blocking:
        print(f"BLOCK: {out_name} {m}", file=sys.stderr)
    for m in provisional:
        print(f"WARN:  {out_name} {m}", file=sys.stderr)

    return {
        "file": out_name, "bid": bid, "cost": cost, "margin": margin,
        "margin_rate": round(mrate*100, 1),
        # ↓ 마진율과 반드시 같이 읽을 것. 보고표의 '가격근거' 컬럼이 이 값이다(규칙 6).
        "price_basis": {
            "est_rows": est_rows,
            "est_amount": est_amount,
            "est_share_of_cost_pct": est_share,
            # 견적을 0원으로 넣은 행 수. >0이면 cost는 하한이므로 검증 완료가 아니다.
            "zero_quote_rows": zero_quote,
            "cost_is_lower_bound": zero_quote > 0,
            "all_verified": est_rows == 0 and zero_quote == 0,
        },
        "margin_rate_is_provisional": est_rows > 0 or zero_quote > 0,
        "reportable": (not blocking) and est_share < 30 and zero_quote == 0,
        "gate": "BLOCKED" if blocking else
                ("PROVISIONAL" if (est_rows or zero_quote) else "OK"),
        "gate_reasons": blocking + provisional,
    }

if __name__ == "__main__":
    data = json.load(open(sys.argv[1], encoding="utf-8"))
    results = []
    for n in data:
        results.append(generate(n["meta"], n["items"], n["out"]))
    print(json.dumps(results, ensure_ascii=False, indent=1))
    blocked = [r for r in results if r["gate"] == "BLOCKED"]
    unrep = [r for r in results if not r["reportable"]]
    if unrep:
        print("\n" + "="*72, file=sys.stderr)
        print("아래 파일은 마진율을 그대로 보고하면 안 된다 (AGENTS.md 규칙 1·6):", file=sys.stderr)
        for r in unrep:
            pb = r["price_basis"]
            why = f"추정 {pb['est_rows']}행 {pb['est_amount']:,}원(원가의 {pb['est_share_of_cost_pct']}%)"
            if pb.get("zero_quote_rows"):
                why += (f" / 견적 0원 계상 {pb['zero_quote_rows']}행 → 원가는 하한, "
                        f"표시 마진율은 상한값일 뿐이다")
            print(f"  {r['gate']:12s} {r['file']} — 마진 {r['margin_rate']}% / {why}", file=sys.stderr)
        print("="*72, file=sys.stderr)
    sys.exit(1 if blocked else 0)
