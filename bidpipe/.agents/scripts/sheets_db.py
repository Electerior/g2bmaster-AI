# -*- coding: utf-8 -*-
"""날짜 폴더(파일 DB) → 구글 스프레드시트(관계형 DB) 어댑터.

왜 옮기나 (2026-08-21 사용자 확정)
  파일 시스템은 PK를 강제하지 못한다. 규칙 4에 "파일명의 공고번호와 B5가 같은지 확인하라"고
  적어놨는데도 대구대 R26DD20847739가 동남보건대 파일의 복사본으로 남아 있었다.
  규칙으로는 안 막히고, 테이블이면 애초에 생기지 않는 사고다.

  append-only(observations.jsonl)도 파일에서는 '규율'일 뿐 강제가 아니다.
  편집기로 줄을 고치면 흔적이 안 남는다. 시트는 셀 편집 이력이 남는다.

계산은 여기서 하지 않는다
  마진·추정가 감사는 audit_prices, 표기는 dashboard_view, 공고 메타는 scan_dashboard를
  **그대로 import**한다. 대시보드·시트·감사가 각자 계산식을 가지면 규칙 6이 갈라진다.

인증
  서비스 계정(.agents/data/credentials/gcp-service-account.json).
  서비스 계정은 Drive 용량이 0이라 **자기 소유 파일을 만들지 못한다**(403 storageQuotaExceeded).
  그래서 문서는 사용자가 만들고 폴더 권한을 상속시킨 것을 쓴다. 새 문서를 API로 만들지 말 것.

사용법
  .agents/py .agents/scripts/sheets_db.py --init      # 시트 6개 생성 + 헤더
  .agents/py .agents/scripts/sheets_db.py --push      # 전체 마이그레이션
  .agents/py .agents/scripts/sheets_db.py --verify    # PK 중복·행수 검증
"""
import os, sys, json, glob, argparse, datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, HERE)

from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

import scan_dashboard
import dashboard_view as dv

KEY_PATH = os.path.join(ROOT, ".agents/data/credentials/gcp-service-account.json")
CONF_PATH = os.path.join(ROOT, ".agents/data/sheets_db.json")
SCOPES = ["https://www.googleapis.com/auth/drive",
          "https://www.googleapis.com/auth/spreadsheets"]

PRICE_DIR = os.path.join(ROOT, ".agents/data/price")
OUTCOMES = os.path.join(ROOT, ".agents/data/outcomes/outcomes.jsonl")

ITEM_START = 26          # gen_analysis.py와 같은 값. 여기서 재정의하지 말 것
STOP_PREFIX = ("합계", "소계", "상품가격", "부대비용", "총액", "총계", "투찰", "원가")


# ─────────────────────────── 스키마 (컬럼 정의는 여기 한 곳만) ───────────────────────────
SCHEMA = {
    "공고": [
        "공고번호", "수요기관", "공고명", "공고일", "마감", "개찰", "납품",
        "입찰방식", "낙찰방법", "참가자격", "이율", "이율근거",
        "추정가격", "기초금액", "투찰가", "원가합계", "부대비용(참고)", "마진액",
        "마진율", "마진표기", "정렬키", "가격근거",
        "선택행수", "미달행수", "추정행수", "추정금액", "추정비중",
        "보고가능", "무결성", "상태", "워크북", "폴더", "공고URL", "개요", "갱신시각",
    ],
    "물품내역": [
        "공고번호", "행", "순번태그", "품명", "요구사양", "수량",
        "판매처", "URL", "단가", "선택", "선택금액", "비고",
    ],
    "상품": [
        "상품키", "품명", "브랜드", "모델코드", "사양", "수량기준",
        "변동성", "최초관측", "URL", "별칭", "출처파일", "메모",
    ],
    "가격관측": [
        "상품키", "관측시각", "관측자", "방법", "티어", "bound", "취득",
        "가격", "통화", "환율", "VAT포함", "판매처", "판매처범위",
        "리스팅", "정품", "URL", "유효기간", "출처파일", "메모",
    ],
    "개찰결과": [
        "공고번호", "상태", "참조번호", "개찰시각", "공고명", "발주기관", "수요기관",
        "예정가격", "기초금액", "낙찰자", "낙찰가", "투찰수", "수집시각", "출처", "메모",
    ],
    "_메타": ["항목", "값", "설명"],
}


# ─────────────────────────── 인증 ───────────────────────────
def services():
    creds = Credentials.from_service_account_file(KEY_PATH, scopes=SCOPES)
    return (build("sheets", "v4", credentials=creds, cache_discovery=False),
            build("drive", "v3", credentials=creds, cache_discovery=False))


def load_conf():
    with open(CONF_PATH, encoding="utf-8") as fh:
        return json.load(fh)


# ─────────────────────────── 값 정규화 ───────────────────────────
def _s(v):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat(sep=" ")[:16] if isinstance(v, dt.datetime) else v.isoformat()
    return str(v)


def _n(v):
    """숫자는 숫자로 넣는다(정렬·집계용). 아니면 빈칸."""
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else ""


def _b(v):
    return "TRUE" if v else ("FALSE" if v is not None else "")


# ─────────────────────────── 물품내역 추출 ───────────────────────────
def read_items(path, notice_no):
    """워크북의 물품 행을 펼친다. 규칙 2 — 묶음 금지이므로 원본 행을 그대로 옮긴다."""
    out = []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return out
    ws = wb.active
    blank = 0
    for r in range(ITEM_START, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name is None or not str(name).strip():
            blank += 1
            if blank >= 5:
                break
            continue
        blank = 0
        nm = str(name).strip()
        if nm.startswith(STOP_PREFIX):
            break
        qty, unit = ws.cell(r, 3).value, ws.cell(r, 6).value
        if _n(qty) == "" and _n(unit) == "":
            continue                      # 소제목·구분선 행
        tag = ""
        if nm.startswith("["):
            end = nm.find("]")
            if end > 0:
                tag = nm[1:end]
        out.append([
            notice_no, r, tag, nm, _s(ws.cell(r, 2).value), _n(qty),
            _s(ws.cell(r, 4).value), _s(ws.cell(r, 5).value), _n(unit),
            _s(ws.cell(r, 7).value), _n(ws.cell(r, 8).value), _s(ws.cell(r, 9).value),
        ])
    wb.close()
    return out


# ─────────────────────────── 행 만들기 ───────────────────────────
def notice_rows(payload):
    rows = []
    for r in payload["rows"]:
        view = r.get("view") or {}
        rows.append([
            _s(r.get("notice_no")), _s(r.get("agency")), _s(r.get("title")),
            _s(r.get("posted_at")), _s(r.get("deadline")), _s(r.get("open_at")),
            _s(r.get("delivery")), _s(r.get("method")), _s(r.get("award")),
            _s(r.get("eligibility")), _n(r.get("rate")), _s(r.get("rate_basis")),
            _n(r.get("est_price")), _n(r.get("base")), _n(r.get("bid")),
            _n(r.get("cost")), _n(r.get("extra_cost")), _n(r.get("margin")),
            _n(r.get("margin_rate")), _s(view.get("text")), _n(view.get("sort")),
            _s(view.get("note")),
            _n(r.get("sel_rows")), _n(r.get("unmet_rows")), _n(r.get("est_rows")),
            _n(r.get("est_amount")), _n(r.get("est_share")),
            _b(r.get("reportable")), _s(r.get("integrity")), _s(r.get("status")),
            _s(r.get("filename")), _s(r.get("folder")), _s(r.get("url")),
            _s(r.get("summary")), _s(r.get("mtime")),
        ])
    return rows


def product_rows():
    p = os.path.join(PRICE_DIR, "products.json")
    if not os.path.exists(p):
        return []
    d = json.load(open(p, encoding="utf-8"))
    rows = []
    for key, v in sorted(d.items()):
        urls = v.get("urls") or []
        rows.append([
            key, _s(v.get("name")), _s(v.get("brand")), _s(v.get("model_code")),
            _s(v.get("spec")), _s(v.get("qty_basis")), _s(v.get("volatility")),
            _s(v.get("first_seen")), urls[0] if urls else "",
            ", ".join(v.get("aliases") or []), _s(v.get("source_file")), _s(v.get("note")),
        ])
    return rows


def observation_rows():
    p = os.path.join(PRICE_DIR, "observations.jsonl")
    if not os.path.exists(p):
        return []
    rows = []
    with open(p, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            o = json.loads(line)
            rows.append([
                _s(o.get("key")), _s(o.get("observed_at")), _s(o.get("observer")),
                _s(o.get("method")), _s(o.get("tier")), _s(o.get("bound")),
                _s(o.get("acquisition")), _n(o.get("price")), _s(o.get("currency")),
                _n(o.get("fx_rate")), _b(o.get("vat_included")), _s(o.get("vendor")),
                _s(o.get("vendor_scope")), _s(o.get("listing_status")),
                _s(o.get("genuine")), _s(o.get("url")), _s(o.get("valid_until")),
                _s(o.get("source_file")), _s(o.get("note")),
            ])
    return rows


def outcome_rows():
    if not os.path.exists(OUTCOMES):
        return []
    rows = []
    with open(OUTCOMES, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            o = json.loads(line)
            bidders = o.get("bidders") or []
            win = bidders[0] if bidders else {}
            rows.append([
                _s(o.get("bid_no")), _s(o.get("status")), _s(o.get("ref_no")),
                _s(o.get("opened_at")), _s(o.get("title")), _s(o.get("org")),
                _s(o.get("demand_org")), _n(o.get("est_price")), _n(o.get("base_price")),
                _s(win.get("name")), _n(win.get("price")), len(bidders),
                _s(o.get("collected_at")), _s(o.get("source")), _s(o.get("note")),
            ])
    return rows


# ─────────────────────────── 시트 조작 ───────────────────────────
def ensure_sheets(svc, doc_id):
    """SCHEMA의 시트가 없으면 만들고, 기본 'Sheet1'은 지운다."""
    meta = svc.spreadsheets().get(spreadsheetId=doc_id).execute()
    have = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta["sheets"]}
    reqs = []
    for name in SCHEMA:
        if name not in have:
            reqs.append({"addSheet": {"properties": {"title": name,
                                                     "gridProperties": {"frozenRowCount": 1}}}})
    if reqs:
        svc.spreadsheets().batchUpdate(spreadsheetId=doc_id, body={"requests": reqs}).execute()
        meta = svc.spreadsheets().get(spreadsheetId=doc_id).execute()
        have = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta["sheets"]}
    for junk in ("시트1", "Sheet1"):
        if junk in have and len(have) > 1:
            svc.spreadsheets().batchUpdate(
                spreadsheetId=doc_id,
                body={"requests": [{"deleteSheet": {"sheetId": have[junk]}}]}).execute()
            have.pop(junk)
    return have


def write_table(svc, doc_id, name, header, rows):
    """헤더 + 데이터를 통째로 갈아끼운다. 값 없는 잔여 행은 지운다."""
    svc.spreadsheets().values().clear(spreadsheetId=doc_id, range=f"'{name}'").execute()
    body = {"values": [header] + rows}
    svc.spreadsheets().values().update(
        spreadsheetId=doc_id, range=f"'{name}'!A1",
        valueInputOption="RAW", body=body).execute()
    return len(rows)


def fill_item_amount(svc, doc_id, n_rows):
    """선택금액을 시트 수식으로 채운다.

    워크북 H열은 `=IF($G26="O",$C26*$F26,0)` 인데, gen_analysis.py(openpyxl)가 만든 파일에는
    **수식 캐시값이 없다**. 이 맥은 Excel·soffice가 전부 막혀 있어 캐시가 생길 일도 없으므로
    data_only=True는 영원히 None을 준다. 그래서 값을 옮기는 대신 같은 수식을 시트에 심는다.
    (F=수량, I=단가, J=선택 → K=선택금액. 원본과 의미가 동일하다)
    """
    if n_rows <= 0:
        return
    vals = [[f'=IF($J{r}="O",$F{r}*$I{r},0)'] for r in range(2, n_rows + 2)]
    svc.spreadsheets().values().update(
        spreadsheetId=doc_id, range=f"'물품내역'!K2",
        valueInputOption="USER_ENTERED", body={"values": vals}).execute()


def format_header(svc, doc_id, sheet_ids):
    reqs = []
    for name, sid in sheet_ids.items():
        if name not in SCHEMA:
            continue
        reqs.append({"repeatCell": {
            "range": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": 1},
            "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True},
                "backgroundColor": {"red": 0.85, "green": 0.91, "blue": 0.83}}},
            "fields": "userEnteredFormat(textFormat,backgroundColor)"}})
        reqs.append({"updateSheetProperties": {
            "properties": {"sheetId": sid, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount"}})
        reqs.append({"setBasicFilter": {"filter": {"range": {"sheetId": sid}}}})
    if reqs:
        svc.spreadsheets().batchUpdate(spreadsheetId=doc_id, body={"requests": reqs}).execute()


# ─────────────────────────── 명령 ───────────────────────────
def cmd_init(svc, doc_id):
    ids = ensure_sheets(svc, doc_id)
    for name, header in SCHEMA.items():
        svc.spreadsheets().values().update(
            spreadsheetId=doc_id, range=f"'{name}'!A1",
            valueInputOption="RAW", body={"values": [header]}).execute()
    format_header(svc, doc_id, ids)
    print("[init] 시트 준비 완료:", ", ".join(SCHEMA))
    return ids


def cmd_push(svc, doc_id):
    ids = ensure_sheets(svc, doc_id)
    payload = scan_dashboard.build_payload()

    n_notice = write_table(svc, doc_id, "공고", SCHEMA["공고"], notice_rows(payload))

    items = []
    for r in payload["rows"]:
        items += read_items(os.path.join(ROOT, r["file"]), _s(r.get("notice_no")))
    n_item = write_table(svc, doc_id, "물품내역", SCHEMA["물품내역"], items)
    fill_item_amount(svc, doc_id, n_item)

    n_prod = write_table(svc, doc_id, "상품", SCHEMA["상품"], product_rows())
    n_obs = write_table(svc, doc_id, "가격관측", SCHEMA["가격관측"], observation_rows())
    n_out = write_table(svc, doc_id, "개찰결과", SCHEMA["개찰결과"], outcome_rows())

    now = dt.datetime.now().isoformat(timespec="seconds")
    meta = [
        ["푸시 시각", now, "이 시각의 파일 상태다"],
        ["공고", n_notice, "PK=공고번호"],
        ["물품내역", n_item, "FK=공고번호. 규칙 2 — 묶음 금지, 원본 행 그대로"],
        ["상품", n_prod, "PK=상품키"],
        ["가격관측", n_obs, "append-only 로그"],
        ["개찰결과", n_out, "규칙 9"],
        ["추정비중 차단선", scan_dashboard.EST_SHARE_BLOCK, "이상이면 마진율 비보고(규칙 6)"],
        ["계산 주체", "audit_prices.py / dashboard_view.py",
         "시트는 계산하지 않는다. 규칙이 갈라지지 않게 하기 위함"],
    ]
    write_table(svc, doc_id, "_메타", SCHEMA["_메타"], [[_s(a), _s(b), _s(c)] for a, b, c in meta])
    format_header(svc, doc_id, ids)

    print(f"[push] 공고 {n_notice} · 물품 {n_item} · 상품 {n_prod} · 관측 {n_obs} · 개찰 {n_out}")
    return {"공고": n_notice, "물품내역": n_item, "상품": n_prod,
            "가격관측": n_obs, "개찰결과": n_out}


def cmd_verify(svc, doc_id):
    """PK 무결성 검사. 파일 시스템이 못 하던 것 = 이걸 하려고 옮겼다."""
    res = svc.spreadsheets().values().batchGet(
        spreadsheetId=doc_id,
        ranges=["'공고'!A2:A", "'물품내역'!A2:A", "'상품'!A2:A", "'가격관측'!A2:A"]).execute()
    vr = res.get("valueRanges", [])

    def col(i):
        return [x[0] for x in (vr[i].get("values") or []) if x and x[0]]

    notices, items, prods, obs = col(0), col(1), col(2), col(3)
    print(f"공고 {len(notices)}행 / 물품 {len(items)}행 / 상품 {len(prods)}행 / 관측 {len(obs)}행")

    ok = True
    from collections import Counter
    dup = {k: v for k, v in Counter(notices).items() if v > 1}
    if dup:
        ok = False
        print("\n[!] 공고번호 PK 중복:", dup)
        print("    → 파일 시스템에서는 이게 안 보였다. 규칙 4 사고가 여기서 드러난다")
    else:
        print("\n[OK] 공고번호 PK 유일")

    orphan = sorted(set(items) - set(notices))
    if orphan:
        ok = False
        print("[!] 고아 물품행(공고 없음):", orphan[:10])
    else:
        print("[OK] 물품내역 FK 무결")

    o_orphan = sorted(set(obs) - set(prods))
    if o_orphan:
        print(f"[!] 미등록 상품키 관측 {len(o_orphan)}건:", o_orphan[:5])
    else:
        print("[OK] 가격관측 FK 무결")
    return ok


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--init", action="store_true")
    ap.add_argument("--push", action="store_true")
    ap.add_argument("--verify", action="store_true")
    a = ap.parse_args()

    conf = load_conf()
    doc_id = conf["doc_id"]
    svc, _drive = services()

    if a.init:
        cmd_init(svc, doc_id)
    if a.push:
        cmd_push(svc, doc_id)
    if a.verify:
        cmd_verify(svc, doc_id)
    if not (a.init or a.push or a.verify):
        ap.print_help()
    print("\n" + conf["url"])


if __name__ == "__main__":
    main()
