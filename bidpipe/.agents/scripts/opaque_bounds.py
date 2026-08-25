# -*- coding: utf-8 -*-
"""개찰결과 역산 — 값을 모르는 품목의 **상한**을 공짜로 얻는다.

원리
    낙찰가 W ≥ Σ(낙찰자 원가)  (마진 ≥ 0 가정)
    우리 BOM에서 공개가로 값을 아는 행의 합을 K라 하면
        Σ(불투명 행) ≤ W − K − 부대비용
    불투명 행이 하나뿐이면 그 행의 단가 상한이 바로 나온다.

    진 공고가 오히려 정보다. 낙찰가가 우리 추정보다 낮으면
    **우리 불투명 항목 추정이 과대했다**는 뜻이고, 그 사실이 다음 견적 협상의 근거가 된다.

한계 (반드시 같이 읽을 것)
    - 낙찰자의 BOM이 우리와 같다는 보장은 없다. 그래서 이건 '가격'이 아니라 '상한'이다.
    - 낙찰자가 역마진 투찰을 했다면 상한이 실제보다 낮게 나온다 → 이상치는 리포트에 남긴다.
    - 관측은 전부 bound=upper 로 기록된다. 원가 행에 그대로 쓰면 게이트가 막는다.

사용:
  python opaque_bounds.py                # 리포트
  python opaque_bounds.py --register     # 가격 DB에 상한 관측으로 기록
"""
import os, re, sys, json, glob, argparse, datetime, collections
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import price_schema as ps
from gen_analysis import url_problem, upper_bound_problem
from outcome_tracker import load_outcomes, latest_by_no, scan_workbooks

ROOT = os.environ.get("BIDPIPE_ROOT") or os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
R0, R1 = 26, 45


def classify_rows(path):
    """워크북의 선택=O 행을 '값을 아는 행'과 '불투명 행'으로 가른다."""
    wb = load_workbook(path, data_only=False)
    ws = wb["분석"] if "분석" in wb.sheetnames else wb.active
    num = lambda v, d=0: v if isinstance(v, (int, float)) else d
    known, opaque = [], []
    for r in range(R0, R1 + 1):
        name = ws.cell(row=r, column=1).value
        if not name or str(ws.cell(row=r, column=7).value or "").strip() != "O":
            continue
        qty = num(ws.cell(row=r, column=3).value, 1)
        price = num(ws.cell(row=r, column=6).value, 0)
        url = str(ws.cell(row=r, column=5).value or "").strip()
        vendor = str(ws.cell(row=r, column=4).value or "")
        note = str(ws.cell(row=r, column=9).value or "")
        row = {"row": r, "name": str(name)[:52], "qty": qty, "price": price,
               "amount": qty * price, "url": url, "vendor": vendor, "note": note,
               "spec": str(ws.cell(row=r, column=2).value or "")[:80]}
        bad = url_problem(url) or upper_bound_problem(url, vendor, note)
        quoted = ("견적" in vendor) or ("견적" in note)
        if bad or quoted:
            row["why"] = bad or "견적 표기"
            opaque.append(row)
        else:
            known.append(row)
    extra = num(ws["B50"].value, 0)
    return known, opaque, extra


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--register", action="store_true")
    ap.add_argument("--assume-margin", type=float, default=0.0,
                    help="낙찰자 최소 마진율 가정(%%). 0이면 순수 상한")
    a = ap.parse_args()

    outs = latest_by_no(load_outcomes())
    if not outs:
        print("개찰결과 로그가 비어 있다 (.agents/data/outcomes/outcomes.jsonl).")
        print("먼저 outcome_tracker.py pending → 수집 → merge 를 돌려라.")
        print("\n[참고] 지금 워크북들의 불투명 행 현황만 먼저 보여준다:\n")

    books = {os.path.basename(w["file"]): w for w in scan_workbooks()}
    rows_out, preview = [], []
    for fname, w in sorted(books.items()):
        path = os.path.join(ROOT, w["file"])
        try:
            known, opaque, extra = classify_rows(path)
        except Exception as e:
            continue
        if not opaque:
            continue
        K = sum(r["amount"] for r in known)
        o_amt = sum(r["amount"] for r in opaque)
        oc = outs.get(w["bid_no"]) if w["bid_no"] else None
        win = (oc or {}).get("win_price") or (oc or {}).get("winning_price")
        preview.append({"file": fname, "known": K, "extra": extra,
                        "opaque_rows": len(opaque), "opaque_amt": o_amt,
                        "win": win, "opaque": opaque})
        if not win:
            continue
        budget = win * (1 - a.assume_margin / 100.0) - K - extra
        rows_out.append({"file": fname, "bid_no": w["bid_no"], "win": win,
                         "known": K, "extra": extra, "budget": budget,
                         "opaque": opaque})

    print("== 불투명 행이 있는 워크북 %d건 ==" % len(preview))
    for p in sorted(preview, key=lambda x: -x["opaque_amt"])[:20]:
        tag = "낙찰 %s원" % format(p["win"], ",") if p["win"] else "개찰결과 미회수"
        print("\n  %s  [%s]" % (p["file"][:60], tag))
        print("     확정 %12s원 + 부대 %10s원 | 불투명 %d행 %s원"
              % (format(p["known"], ","), format(p["extra"], ","),
                 p["opaque_rows"], format(p["opaque_amt"], ",")))
        for r in p["opaque"][:6]:
            print("       · %-46s %2d개 %11s원  (%s)"
                  % (r["name"][:46], r["qty"], format(r["price"], ","), r["why"][:28]))

    if not rows_out:
        print("\n낙찰가가 회수된 건이 없어 상한을 계산할 수 없다.")
        return

    print("\n== 개찰 역산 상한 ==")
    products, recs = ps.load_products(), []
    today = datetime.date.today().isoformat()
    for x in rows_out:
        n = len(x["opaque"])
        print("\n[%s] 낙찰 %s원 − 확정 %s원 − 부대 %s원 = 불투명 %d행 예산 %s원"
              % (x["bid_no"], format(x["win"], ","), format(x["known"], ","),
                 format(x["extra"], ","), n, format(int(x["budget"]), ",")))
        if x["budget"] < 0:
            print("   ⚠ 음수다. 낙찰자가 역마진이거나 우리 확정가가 과대하다 → 원가 재검토 대상")
            continue
        for r in x["opaque"]:
            if n == 1 and r["qty"]:
                ub = int(x["budget"] / r["qty"])
                print("   · %-46s 단가 상한 %11s원 (우리 산정 %s원, %+.0f%%)"
                      % (r["name"][:46], format(ub, ","), format(r["price"], ","),
                         (ub - r["price"]) / r["price"] * 100 if r["price"] else 0))
                if a.register:
                    key = "opaque:%s/%s" % (x["bid_no"], ps._slug(r["name"])[:24])
                    products.setdefault(key, {
                        "name": r["name"], "brand": None, "model_code": None,
                        "spec": r["spec"], "urls": [], "aliases": [],
                        "qty_basis": "개당", "volatility": "normal", "first_seen": today,
                        "note": "개찰 역산 상한. %s 낙찰 %s원에서 확정가 차감" % (
                            x["bid_no"], format(x["win"], ",")),
                    })
                    recs.append(ps.observe(
                        key, ub, vendor="개찰 역산(상한)", method="outcome",
                        url=None, observer="script:opaque_bounds",
                        acquisition="derived", bound="upper",
                        note="%s 낙찰 %s / 확정 %s / 부대 %s" % (
                            x["bid_no"], x["win"], x["known"], x["extra"])))
            else:
                print("   · %-46s (공동 제약: %d행이 %s원을 나눠 씀)"
                      % (r["name"][:46], n, format(int(x["budget"]), ",")))

    if a.register and recs:
        ps.save_products(products)
        ps.append_obs(recs)
        ps.rebuild_view()
        print("\n기록: 상한 관측 %d건 (bound=upper)" % len(recs))


if __name__ == "__main__":
    main()
