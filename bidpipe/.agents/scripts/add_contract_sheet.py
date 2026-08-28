#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
분석 워크북에 '계약분석' 시트를 추가한다.
- 기존 '분석' 시트는 절대 건드리지 않는다.
- 서식 팔레트는 Format.xlsx 계열(맑은 고딕 / 2F5597 헤더 / D9E2F3 표머리 / FFF2CC 가정값)을 따른다.
- 저장 직전에 파일을 다시 읽어 사용자 병행 편집을 덮어쓰지 않는다.

usage: python add_contract_sheet.py <xlsx경로>
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SHEET = "계약분석"
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F = "맑은 고딕"
C_HDR_BG = "FF2F5597"      # 섹션 헤더 배경
C_TBL_BG = "FFD9E2F3"      # 표 머리 배경
C_LBL_BG = "FFF2F2F2"      # 라벨 배경
C_WARN_BG = "FFFCE4D6"     # 주의 강조
C_BAD_BG = "FFFFC7CE"      # 위험
C_OK_BG = "FFE2EFDA"       # 양호
C_ASSUM_BG = "FFFFF2CC"    # 가정값
C_GRAY = "FF808080"
C_BLUE = "FF0000FF"

COLS = {"A": 15, "B": 46, "C": 30, "D": 9, "E": 15, "F": 40}


def style(ws, coord, *, v=None, bold=False, size=10, color=None, bg=None,
          wrap=True, ha="left", va="center", border=True, fmt=None):
    c = ws[coord]
    if v is not None:
        c.value = v
    c.font = Font(name=F, size=size, bold=bold, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if border:
        c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c


def section(ws, row, text):
    ws.merge_cells(f"A{row}:F{row}")
    style(ws, f"A{row}", v=text, bold=True, size=11, color="FFFFFFFF",
          bg=C_HDR_BG, ha="left", wrap=False)
    ws.row_dimensions[row].height = 20
    return row + 1


def table_head(ws, row, headers):
    """headers: [(col_letter_or_range, text), ...]"""
    for ref, text in headers:
        if ":" in ref:
            ws.merge_cells(f"{ref[0]}{row}:{ref[-1]}{row}")
            ref = ref[0]
        style(ws, f"{ref}{row}", v=text, bold=True, bg=C_TBL_BG, ha="center", wrap=False)
    ws.row_dimensions[row].height = 18
    return row + 1


def row_cells(ws, row, cells, *, height=None, bg=None):
    """cells: [(ref, value, opts_dict), ...] ref may be 'A' or 'B:D'"""
    used = set()
    for ref, val, opts in cells:
        o = dict(opts or {})
        if bg and "bg" not in o:
            o["bg"] = bg
        if ":" in ref:
            start, end = ref.split(":")
            ws.merge_cells(f"{start}{row}:{end}{row}")
            for ch in range(ord(start), ord(end) + 1):
                used.add(chr(ch))
                if chr(ch) != start:
                    style(ws, f"{chr(ch)}{row}", bg=o.get("bg"))
            ref = start
        else:
            used.add(ref)
        style(ws, f"{ref}{row}", v=val, **o)
    for ch in "ABCDEF":
        if ch not in used:
            style(ws, f"{ch}{row}", bg=bg)
    if height:
        ws.row_dimensions[row].height = height
    autofit(ws, row, [(ref, (opts or {}).get("size", 10)) for ref, _, opts in cells])
    return row + 1


def _needed_height(text, width_units, size):
    """병합 셀은 Excel 자동맞춤이 안 되므로 필요 행높이를 직접 추정한다."""
    if text is None:
        return 0
    s = str(text)
    px_w = width_units * 7.6
    font_px = size * 1.16
    # 한글 1자 ≈ font_px, 영숫자·공백 ≈ 0.55 font_px
    def wide(ch):
        return 1.0 if ord(ch) > 0x1100 else 0.55
    cap = px_w / font_px
    lines = 0
    for para in s.split("\n"):
        if not para:
            lines += 1
            continue
        w = 0.0
        n = 1
        for ch in para:
            w += wide(ch)
            if w > cap:
                n += 1
                w = wide(ch)
        lines += n
    return lines * font_px * 1.32 + 4


def autofit(ws, row, spans):
    """spans: [(ref, size), ...] — 행 내 셀들의 필요 높이 중 최대값으로 행높이를 보정"""
    need = 0
    for ref, size in spans:
        if ":" in ref:
            a, b = ref.split(":")
            wu = sum(COLS[chr(c)] for c in range(ord(a), ord(b) + 1))
            col = a
        else:
            wu = COLS[ref]
            col = ref
        need = max(need, _needed_height(ws[f"{col}{row}"].value, wu, size))
    cur = ws.row_dimensions[row].height or 0
    if need > cur:
        ws.row_dimensions[row].height = round(need, 1)


def note(ws, row, text):
    ws.merge_cells(f"A{row}:F{row}")
    style(ws, f"A{row}", v=text, size=9, color=C_GRAY, border=False, ha="left")
    ws.row_dimensions[row].height = 15
    return row + 1


def build(path):
    wb = openpyxl.load_workbook(path)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    idx = wb.sheetnames.index("분석") + 1 if "분석" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet(SHEET, idx)
    for col, w in COLS.items():
        ws.column_dimensions[col].width = w

    r = 1
    ws.merge_cells("A1:F1")
    style(ws, "A1", v="계약 조건 분석 — 대구대학교 2026학년도 기계기구(개인용컴퓨터·모니터·소형복합기) 단가구매",
          bold=True, size=14, border=False, ha="left", wrap=False)
    ws.row_dimensions[1].height = 28
    r = 2
    ws.merge_cells("A2:F2")
    style(ws, "A2",
          v="공고번호 R26BK01685839 / 대구대학교 입찰공고 제2026-109호 · 조사기준일 2026-08-20 · "
            "1차 출처: 입찰공고문(PDF)·제안요청서(HWP)·나라장터·대구대 전자입찰 낙찰공고·제조사 공식 스펙",
          size=9, color=C_GRAY, border=False, ha="left", wrap=False)
    ws.row_dimensions[2].height = 15

    # ------------------------------------------------------------------ 1
    r = 4
    r = section(ws, r, "1. 계약 개요")
    r = table_head(ws, r, [("A", "항목"), ("B:D", "내용"), ("E:F", "근거")])
    for item, content, src, h in [
        ("계약방식", "일반경쟁 / 입찰가격 및 제안서 심사(협상에 의한 계약) / 현장입찰(직찰)", "공고문 5", 16),
        ("낙찰자 결정", "품목별 유효 입찰자 중 심사로 협상적격자를 선정하고, 고득점자순 협상으로 낙찰자 결정", "공고문 6-가", 30),
        ("품목 그룹", "품목1(1-1~1-4 PC·모니터) 1,517,900,000원 / 품목2(2-1~2-2 복합기) 36,500,000원. "
                    "품목 그룹별로 별도 입찰", "공고문 1-나 / 제안요청서 7-가", 30),
        ("계약기간", "계약체결일 ~ 2027-08-31 (단가계약)", "공고문 1-다", 16),
        ("적용 규정", "「대구대학교 재무·회계규정」 — 입찰 성립·무효, 지체상금 산정에 적용. "
                    "국가계약법이 아니므로 이의신청·분쟁조정 창구 없음", "공고문 7, 15-마 / 제안요청서 5-다", 30),
        ("계약체결 기한", "본교의 계약체결 요청일로부터 3일 이내", "공고문 9-라", 16),
        ("제안서 파일 제출", "이메일 pearl@daegu.ac.kr (2026-09-01 14:00까지 유효)", "공고문 8 ※", 16),
        ("문의처", "규격: 자산관리팀 053-850-5371 / 입찰: 총무팀 053-850-5342", "공고문 15-사", 16),
    ]:
        r = row_cells(ws, r, [
            ("A", item, dict(bold=True, bg=C_LBL_BG)),
            ("B:D", content, {}),
            ("E:F", src, dict(size=9, color=C_GRAY)),
        ], height=h)

    # ------------------------------------------------------------------ 2
    r += 1
    r = section(ws, r, "2. 낙찰자 결정 구조 (배점 100점)")
    r = table_head(ws, r, [("A", "평가영역"), ("B", "평가항목"), ("C", "배점"),
                           ("D", "평가방식"), ("E", "우리 예상"), ("F", "근거 / 비고")])
    for area, item, pt, how, ours, src, h in [
        ("기술능력평가\n(60점)", "경영상태(기업신용평가등급)", 5, "정량",
         "확인필요", "AAA·AA=5 / A~BBB=4 / BB=3 / B=2 / CCC이하=0. 등급확인서 미제출 시 최저등급. [제안요청서 라-2)-가)]", 32),
        ("", "물품규격", 40, "평가위원",
         "30 (부족1 가정)", "동등=35 / 우수 1~5개↑=36~40 / 부족=0~30(1사양당 -5). [라-2)-나)]", 32),
        ("", "수행계획 및 기타 부가제안", 15, "평가위원",
         "13 (우수 가정)", "매우우수 14~15 / 우수 12~13 / 보통 10~11 / 미흡 8~9. [라-2)-다)]", 32),
        ("입찰가격평가\n(40점)", "입찰금액", 40, "상대평가",
         "투찰율 종속", "제안요청서에 산식이 '배점×'까지만 기재되어 원문 확인 불가 → 현설 질문 대상. [라-2)-라)]", 32),
    ]:
        r = row_cells(ws, r, [
            ("A", area, dict(bold=True, bg=C_LBL_BG, ha="center")),
            ("B", item, {}),
            ("C", pt, dict(ha="center", fmt="0")),
            ("D", how, dict(ha="center")),
            ("E", ours, dict(ha="center", bg=C_ASSUM_BG)),
            ("F", src, dict(size=9)),
        ], height=h)
    r = row_cells(ws, r, [
        ("A", "자동탈락선", dict(bold=True, bg=C_BAD_BG, ha="center")),
        ("B:F", "기술능력평가 합계가 배점한도(60점)의 80% = 48점 미만이면 입찰가격 평가를 적용하지 않고 자동 탈락. "
                "→ 가격을 아무리 낮게 써도 봉투가 열리지 않는다. [제안요청서 나-2)]",
         dict(bold=True, bg=C_BAD_BG)),
    ], height=30)
    r = row_cells(ws, r, [
        ("A", "절차 리스크", dict(bold=True, bg=C_WARN_BG, ha="center")),
        ("B:F", "① 필요 서류 미첨부·불명확 시 해당 항목 0점 [나-1)]  "
                "② 심사위원 명단·심사결과 세부내용 비공개, 이의제기 불가 [나-3)]  "
                "③ 평가 결과 비공개, 이의제기 불가 [마-2)-아)]  "
                "④ 제안서·계약서 해석상 이견 시 대구대학교 해석 우선 [마-3)-라)]  "
                "⑤ 본교 상대 소송 이력(소송 중 포함) 시 입찰참가 불가 [공고문 4-다]",
         dict(bg=C_WARN_BG)),
    ], height=52)
    r = note(ws, r, "※ 기술능력 60점 중 55점(물품규격 40 + 수행계획 15)이 평가위원 주관 영역이며 결과가 공개되지 않는다.")

    # ------------------------------------------------------------------ 3
    r += 1
    r = section(ws, r, "3-1. 규격 전 행 대조 — 모니터 3종 (RFP 11개 행 전수)")
    r = table_head(ws, r, [("A", "규격 행"), ("B", "24\" RFP / 후보"), ("C", "32\" RFP / 후보"),
                           ("D:E", "34\" RFP / 후보"), ("F", "판정")])
    mon_rows = [
        ("크기", "60.4cm / LG 24BA450 60.4cm", "86.4cm / 크로스오버 81.3cm",
         "86.4cm / 크로스오버 86.42cm", "32\"는 RFP 오타(32\"=81.3cm). 24\"·34\" 정확 일치", "warn"),
        ("해상도", "1920x1080 / 동일", "2560x1440 / 동일", "3440x1440 / 동일", "충족", "ok"),
        ("명암비", "1000:1 / 1,300:1", "1000:1 / 1,000:1", "3000:1 / 3,000:1", "충족 (24\" 우수)", "ok"),
        ("밝기", "250cd / 250nits", "350cd / 350nits", "350cd / 400nits", "충족 (34\" 우수)", "ok"),
        ("응답속도", "5ms(MPRT 1ms) / 5ms GTG", "1ms / 1ms GTG", "5ms / 1ms OD",
         "24\" MPRT 1ms 표기 없음 = 확인필요", "warn"),
        ("패널", "IPS / IPS", "IPS / IPS", "VA / VA", "충족", "ok"),
        ("입력단자", "2 HDMI + D-Sub / HDMI 1 + DP + D-Sub", "1 DP + 1 HDMI / HDMI 2 + DP 2",
         "1 DP + 2 HDMI / HDMI 2 + DP 2", "24\" HDMI 개수 부족 (국내 대안 없음)", "bad"),
        ("시야각", "178/178 / 광시야각", "178/178 / 광시야각", "178/178 / 광시야각",
         "3종 모두 수치 미공개 = 확인필요", "warn"),
        ("소비전력", "14.9W / LG 10W (Dell 28.5W)", "48.0W / 34W", "42.0W / 26W (G34WQC 80W)",
         "교체 후보는 전부 우수. 기존 선택품 2종은 미달", "bad"),
        ("대기전력", "0.5W / 0.5W", "0.3W / 0.5W", "0.5W / 0.5W 이하",
         "32\" 대기전력 초과 = 부족 (시장 전반 0.5W)", "bad"),
        ("스탠드", "HAS 풀 / 풀 HAS", "HAS 풀 / 풀 HAS", "일반형 / 엘리베이션+틸트+스위블", "충족", "ok"),
    ]
    bgmap = {"bad": C_BAD_BG, "warn": C_WARN_BG, "ok": C_OK_BG}
    for row_name, c24, c32, c34, verdict, tone in mon_rows:
        r = row_cells(ws, r, [
            ("A", row_name, dict(bold=True, bg=C_LBL_BG, ha="center")),
            ("B", c24, dict(size=9)),
            ("C", c32, dict(size=9)),
            ("D:E", c34, dict(size=9)),
            ("F", verdict, dict(size=9, bg=bgmap[tone])),
        ], height=26)
    r = note(ws, r, "후보: 24\" LG 24BA450 197,000 / 32\" 크로스오버 32QD200GM 268,910 / 34\" 크로스오버 34QW290GM 299,000. "
                    "출처: 다나와 상세 스펙시트 2026-08-20 조회.")
    r = row_cells(ws, r, [
        ("A", "모니터 교체 효과", dict(bold=True, bg=C_OK_BG, ha="center")),
        ("B:F", "34\" GIGABYTE G34WQC(396,270, 소비전력 80W 미달) → 크로스오버 34QW290GM(299,000, 전 행 충족·4항목 우수): "
                "대당 -97,270 × 180대 = -17,508,600원. "
                "24\" Dell E2425HSM(176,000, 28.5W 미달) → LG 24BA450(197,000, 10W): 대당 +21,000 × 250대 = +5,250,000원. "
                "순 원가 -12,258,600원 + 부족 사양 2건 제거. → 원가가 줄면서 점수는 오르는 교체다.",
         dict(bg=C_OK_BG)),
    ], height=46)

    r += 1
    r = section(ws, r, "3-2. 규격 적합성 시뮬레이션 (통과선 = 기술능력 48점)")
    r = table_head(ws, r, [("A", "시나리오"), ("B", "구성"), ("C", "부족 사양"),
                           ("D", "물품규격"), ("E", "기술합계"), ("F", "판정 / 원가증감")])
    sims = [
        ("S0 워크북 원안", "ASUS B860M-A-CSM + Dell E2425HSM + 32QD200GM + G34WQC",
         "PC 4건(HDMI·RGB·PS/2·전면USB-C) + 24\" 2건(입력단자·전력) + 32\" 1건(대기전력) + 34\" 1건(전력) = 8건",
         0, 18, "탈락 / 기준", "bad"),
        ("S1 제품 최적화\n(선정의 한계)", "ASRock B860M-HDVP + GT1030 / LG 24BA450 / 32QD200GM / 34QW290GM",
         "① PC 전면 USB-C  ② 24\" 입력단자(HDMI 개수)  ③ 32\" 대기전력 = 3건",
         20, 38, "탈락 / +28,223,400원", "bad"),
        ("S2 현설 완화 1건", "S1 구성 유지 + 24\" '2 HDMI, D-Sub' 합산 해석 인정",
         "① PC 전면 USB-C  ② 32\" 대기전력 = 2건", 25, 43, "탈락 / 동일", "bad"),
        ("S3 현설 완화 2건", "S2 + 32\" 대기전력 0.3W → 0.5W 완화(시장 전반이 0.5W)",
         "① PC 전면 USB-C = 1건", 30, 48, "경계 통과 / 동일", "warn"),
        ("S4 현설 완화 3건", "S3 + 전면 USB-C 완화(후면 USB-C 갈음 또는 삭제)",
         "없음 (제시규격과 동등)", 35, 53, "통과 / 동일", "ok"),
        ("S5 우수 제안", "S4 + 우수 5개 이상(명암비·밝기·응답속도·전력·포트) + 수행계획 만점",
         "없음 (우수 5개 이상)", 40, 60, "통과(여유) / 동일", "ok"),
    ]
    for name, conf, lack, spec_pt, tot, verdict, tone in sims:
        r = row_cells(ws, r, [
            ("A", name, dict(bold=True, ha="center")),
            ("B", conf, dict(size=9)),
            ("C", lack, dict(size=9)),
            ("D", spec_pt, dict(ha="center", fmt="0")),
            ("E", tot, dict(ha="center", bold=True, fmt="0")),
            ("F", verdict, dict(ha="center", bold=True)),
        ], height=34, bg=bgmap[tone])
    r = note(ws, r, "산식: 기술합계 = 경영상태 5(AA 이상 가정) + 물품규격 + 수행계획 13(우수 가정). "
                    "물품규격 부족 구간은 30점에서 시작해 1사양당 5점 감점, 하한 0점 [제안요청서 라-2)-나)].")
    r = note(ws, r, "원가증감은 워크북 원안 대비값: PC +67,470×600 = +40,482,000 / 24\" +21,000×250 = +5,250,000 / "
                    "34\" -97,270×180 = -17,508,600 → 순 +28,223,400원.")
    r = note(ws, r, "※ 경영상태가 4점(A~BBB)으로 내려가면 S4는 47점이 되어 탈락한다 → 신용평가등급확인서 사전 확보가 통과 조건.")
    r = row_cells(ws, r, [
        ("A", "핵심 결론", dict(bold=True, bg=C_BAD_BG, ha="center")),
        ("B:F", "시중 제품을 최적으로 골라도 부족 3건(38점)이 한계이며 이는 자동탈락 구간이다. "
                "통과선 48점에 닿으려면 8/26 현설에서 최소 2건을 완화받아야 한다. "
                "→ 현설 결과가 투찰 여부의 유일한 선결 조건이고, 제품 교체는 그다음 문제다.",
         dict(bold=True, bg=C_BAD_BG)),
    ], height=34)
    r = row_cells(ws, r, [
        ("A", "시장 제약\n(실측)", dict(bold=True, bg=C_WARN_BG, ha="center")),
        ("B:F", "① 네이티브 HDMI 2포트 B860 보드는 존재하지 않음(B860 라인업 전수조사 완료)  "
                "② D-Sub(RGB) 탑재 B860 보드는 ASRock B860M-HDVP 1종뿐이고 유통사도 (주)에즈윈 단독  "
                "③ 이 보드에는 전면 Type-C 헤더가 없고, 다나와 검색 기준 전면 베이용 USB-C 패널 제품이 없음. "
                "후면 확장 브라켓(라이트컴 COMS TB803, 22,930원)은 후면 포트라 '전면' 요구를 충족하지 못함  "
                "④ 24\" 모니터 '2 HDMI + D-Sub + HAS 피봇' 전조건 충족 신제품 부재(LG 24MP500 단종, Dell E2424H 피봇 없음)  "
                "⑤ 32\" 대기전력 0.3W 이하 제품 부재 — 크로스오버 32QD200GM·알파스캔 32Q50G 모두 0.5W",
         dict(bg=C_WARN_BG)),
    ], height=70)

    # ------------------------------------------------------------------ 4
    r += 1
    r = section(ws, r, "4. 단가계약 조항 (수량·발주)")
    r = table_head(ws, r, [("A", "조항"), ("B:C", "원문 요지"), ("D:E", "우리 영향"), ("F", "근거")])
    for item, txt, impact, src, h in [
        ("수량 하한 없음", "최종 납품수량이 구매예정수량에 미달하거나 초과하여도 이의를 제기할 수 없음. "
                       "구매 예정수량은 대학 학사 상황에 따라 가감될 수 있음",
         "1,517,900,000원은 매출이 아니라 상한선. 발주 60% 시 마진 2.33억→1.40억 (마진율은 유지, 금액만 감소)",
         "제안요청서 2 ※, 4-나 / 공고문 15-라", 46),
        ("초과분 단가 고정", "계약기간 내 예정수량을 초과할 경우 계약금액으로 제품을 계속 공급하여야 함",
         "부품가 상승 시 상한 없는 역마진 노출. RAM이 PC 원가의 약 25%(대당 353,400원)",
         "공고문 15-라", 32),
        ("발주·설치", "대학 요청 시 품목별 단가를 적용해 지정 일시·장소에 납품 및 설치. 세부는 별도로 정함",
         "1년간 수시 분산 납품(경산캠퍼스 + 대구 분관) 물류·설치비가 반복 발생",
         "제안요청서 5-가", 32),
        ("납기 연기", "납품기한 3일 전까지 사유 명시 연기신청서 제출 + 승인받은 경우에만 지체상금 면제",
         "승인 전 지연은 지체상금. 부품 수급 지연 시 사전 신청 프로세스 필요",
         "제안요청서 5-나", 32),
        ("지체상금", "금액·비율은 대구대학교 재무·회계규정에 의하여 산정 (규정 원문 비공개)",
         "요율을 모른 채 계약하게 됨 → 현설에서 확인 필요",
         "제안요청서 5-다", 32),
    ]:
        r = row_cells(ws, r, [
            ("A", item, dict(bold=True, bg=C_LBL_BG, ha="center")),
            ("B:C", txt, {}),
            ("D:E", impact, {}),
            ("F", src, dict(size=9, color=C_GRAY)),
        ], height=h)

    # ------------------------------------------------------------------ 5
    r += 1
    r = section(ws, r, "5. 보증금·위약벌·하자보수")
    r = table_head(ws, r, [("A", "구분"), ("B", "금액 / 요율"), ("C:D", "조건"),
                           ("E", "예상액(원)"), ("F", "리스크 / 근거")])
    for item, rate, cond, amt, risk, h in [
        ("입찰보증금", "품목별 입찰금액의 5/100 이상\n이행(입찰)보증보험증권",
         "보증기간: 입찰일 이전 ~ 입찰일 30일 이후. 면제조항을 본 입찰에 적용하지 아니함",
         64510750,
         "품목1 투찰 85%(1,290,215,000원) 기준 5% 산정. 보증채권자 대구대학교 515-82-02603 [공고문 9]", 40),
        ("낙찰 후 미계약", "입찰보증금 전액 귀속",
         "계약체결 요청일로부터 3일 이내 미체결 시 위약벌로 전액 귀속하고, 위약벌과 별개로 손해배상 청구 가능",
         None,
         "국가계약법의 '손해배상 예정'과 달리 위약벌 + 손해배상 별도 = 배상 상한 없음 [공고문 9-라]", 40),
        ("계약보증금", "계약금액의 10/100\n이행(계약)보증보험증권", "계약체결 시 제출",
         None, "[공고문 14-가]", 30),
        ("계약 해지·해제", "계약보증금 전액 귀속",
         "계약상대자 사유로 해지·해제 시 시점과 관계없이 위약벌로 전액 귀속 + 손해배상 별도 청구",
         None, "[공고문 14-나]", 40),
        ("하자보수보증금", "실제 발주 총금액의 10/100\n이행(하자)보증보험증권",
         "최종 납품 완료 시 제출. 무상 하자보증은 계약 종료일(2027-08-31)로부터 1년",
         None,
         "실질 A/S 기간 최대 2년(2026-09 첫 납품분 기준). 하자 여부는 사용 부서·검수부서가 최종 결정 [제안요청서 6]", 40),
    ]:
        r = row_cells(ws, r, [
            ("A", item, dict(bold=True, bg=C_LBL_BG, ha="center")),
            ("B", rate, dict(ha="center", size=9)),
            ("C:D", cond, {}),
            ("E", amt, dict(ha="center", fmt="#,##0")),
            ("F", risk, dict(size=9)),
        ], height=h)

    # ------------------------------------------------------------------ 6
    r += 1
    r = section(ws, r, "6. 일정 및 구비서류")
    r = table_head(ws, r, [("A", "일시"), ("B:C", "항목"), ("D", "필수"), ("E:F", "비고 / 근거")])
    for when, what, must, memo, h in [
        ("2026-08-26(수)\n15:00", "과업(현장)설명 — 경산캠퍼스 성산홀(본관) 10층 입찰실",
         "필수", "미참가 시 입찰참가 자격 없음. 제출: 과업설명 참가신청서·사업자등록증 사본·개인정보 수집이용 동의서·"
                "청렴계약이행 서약서 / 지참: 입찰공고문·제안요청서 [공고문 2, 3, 4-가]", 46),
        ("2026-09-01(화)\n13:40~14:00", "입찰등록 (현장)", "필수", "시간 엄수. 제출 서류 일체 반환 불가 [공고문 2, 15-다]", 30),
        ("2026-09-01(화)\n14:00", "현장 투찰. 개찰은 제안서 심사 시 시행", "필수",
         "입찰서는 품목별 1인 1통, 산출내역서 포함 밀봉. 금액은 한글·아라비아숫자 병기(불일치 시 한글 우선). "
         "제출 후 교환·변경·취소 불가 [공고문 10, 11]", 46),
        ("등록 시 구비서류", "입찰참가신청서(법인인감 날인) / 사업자등록증 사본 / 법인등기사항 전부증명서 / "
                        "인감증명서·사용인감계 / 재직증명서·위임장 / 품목별 이행(입찰)보증보험증권 / "
                        "제안서 8부 + 저장매체 / 개인정보 동의서 / 청렴계약 서약서 / 서약서 / 사용인감 지참",
         "필수", "등기부·인감증명서는 입찰일 기준 3개월 이내 발급 원본. 사본은 원본대조필. 제안서 제본 금지 [공고문 8]", 60),
        ("제안서", "A4 20장 내외 8부. 목차 고정: Ⅰ.제안개요 / Ⅱ.물품목록 / Ⅲ.사업지원부문. "
                 "서식 제2호 [상세내역서] 품목별 필수 작성",
         "필수", "모호한 표현(\"~할 수도 있다\")은 \"할 수 없다\"로 간주. 제안 내용은 계약서와 동일한 효력 "
                "[제안요청서 마-1), 2), 3)]", 46),
        ("신용평가등급확인서", "회사채·기업어음 또는 기업신용평가등급. 공고일 전일 이전 평가 + 유효기간 내",
         "필수", "미제출 시 경영상태 최저등급 → 기술점수 -5점 → 자동탈락 직결 [제안요청서 라-2)-가)]", 40),
    ]:
        tone = C_WARN_BG if must == "필수" else None
        r = row_cells(ws, r, [
            ("A", when, dict(bold=True, bg=C_LBL_BG, ha="center", size=9)),
            ("B:C", what, {}),
            ("D", must, dict(ha="center", bold=True, bg=tone)),
            ("E:F", memo, dict(size=9)),
        ], height=h)

    # ------------------------------------------------------------------ 7
    r += 1
    r = section(ws, r, "7. 과거 이력 및 경쟁 구도")
    r = table_head(ws, r, [("A", "연도"), ("B", "공고번호"), ("C", "추정가격(원)"),
                           ("D:E", "개찰결과 공개 여부"), ("F", "비고")])
    for y, no, price, res, memo in [
        ("2023", "20230731033", 352818182, "유찰 — 사유 \"입찰 미성립\"", "직찰"),
        ("2023", "20230807670", 20090909, "유찰 — 사유 \"직찰로 진행함\"", "재입찰 성격"),
        ("2024", "20240802497", 531909091, "유찰 — 사유 \"입찰 미성립\"", "직찰"),
        ("2024", "20240814470", None, "유찰 — 사유 \"수요기관 직찰로 진행\"", "추정가격 미확인"),
        ("2025", "R25BK01003491", 1170245455, "개찰결과분류조회 \"데이터가 없음\"", "멀티미디어 포함"),
        ("2025", "R25BK01014865", 144018182, "개찰결과분류조회 \"데이터가 없음\"", "재공고"),
        ("2026", "R26BK01685839", 1413090909, "개찰 전 (2026-09-01)", "본 건"),
    ]:
        r = row_cells(ws, r, [
            ("A", y, dict(ha="center")),
            ("B", no, dict(ha="center")),
            ("C", price, dict(ha="right", fmt="#,##0")),
            ("D:E", res, {}),
            ("F", memo, dict(size=9)),
        ], height=16)
    r = row_cells(ws, r, [
        ("A", "결론", dict(bold=True, bg=C_WARN_BG, ha="center")),
        ("B:F", "동일 계열 공고가 2023년부터 매년 반복되고 추정가격은 3.53억→5.32억→11.70억→14.13억으로 증가했다. "
                "그러나 전 연도가 '직찰'(현장입찰)로 등록되어 낙찰업체·낙찰금액이 나라장터에 남지 않았다 "
                "→ 경쟁사 투찰 수준을 역산할 근거가 존재하지 않는다. "
                "확인 경로: 나라장터 입찰공고 검색(수요기관 대구대학교, 기관코드 7001607) + 입찰개찰/낙찰 → 개찰결과분류조회.",
         dict(bg=C_WARN_BG)),
    ], height=46)

    r += 1
    r = table_head(ws, r, [("A", "품목 키워드"), ("B", "2024"), ("C", "2025"),
                           ("D:E", "2026(8월까지)"), ("F", "출처")])
    for kw, y24, y25, y26, src in [
        ("복합기", "삼성프라자 7/8건 (88%)", "삼성프라자 20/22건 (91%)", "삼성프라자 7/8건 (88%)",
         "bid.daegu.ac.kr → 낙찰공고 → 소액견적"),
        ("컴퓨터", "디지탈맥스 27/42건 (64%)", "디지탈맥스 21/58건 (36%)", "디지탈맥스 6/18건 (33%)",
         "동일 (학년도 + 견적건명 검색)"),
        ("모니터", "디지탈맥스 다수", "태성미디어 6/23건(26%), 디지탈맥스 4/23건(17%)", "표본 부족(낙찰 2건)",
         "동일"),
    ]:
        r = row_cells(ws, r, [
            ("A", kw, dict(bold=True, bg=C_LBL_BG, ha="center")),
            ("B", y24, dict(ha="center", size=9)),
            ("C", y25, dict(ha="center", size=9)),
            ("D:E", y26, dict(ha="center", size=9)),
            ("F", src, dict(size=9, color=C_GRAY)),
        ], height=18)
    r = note(ws, r, "※ 위 점유율은 대구대 '소액견적' 낙찰 기준이며, 본 단가계약의 낙찰자와 동일하다는 근거는 없다(정황 추정). "
                    "유찰·공고취소 건은 모수에서 제외했다.")

    r += 1
    r = row_cells(ws, r, [
        ("A", "규격서 노후 징후", dict(bold=True, bg=C_LBL_BG, ha="center")),
        ("B:F", "① \"Internet Explorer 11\" — 2022-06 지원종료된 브라우저가 OS/브라우저 항목에 잔존  "
                "② FDD·ODD 행 잔존  "
                "③ \"Reaitek\"(Realtek 오타), \"Enternet\"(Ethernet 오타), \"2.5GHz Gigabit\"(2.5GbE 오기)  "
                "④ 32인치 크기 \"86.4cm\" (32\"=81.3cm, 86.4cm는 34\" 값)  "
                "⑤ I/O \"1 PS/2, 2 HDMI, 1 RGB\" — 국내 유통 B860 보드 중 동시 충족 제품 없음  "
                "→ 규격표가 수년간 복사되어 왔고 특정 제품의 사양을 전재했을 개연성이 높다. "
                "AGENTS.md 규칙 10(규격 제약 민감도 분석) 적용 대상.",
         {}),
    ], height=76)

    # ------------------------------------------------------------------ 8
    r += 1
    r = section(ws, r, "8. 리스크 등록부")
    r = table_head(ws, r, [("A", "ID"), ("B", "리스크"), ("C", "발생 조건"),
                           ("D", "등급"), ("E", "영향(원)"), ("F", "대응")])
    risks = [
        ("R1", "규격평가 자동탈락 (기술능력 48점 미만)", "현설 완화 없이 투찰 (제품 선정 최고 43점)", "치명", 233028600,
         "8/26 현설에서 부족 판정 단위 + 24\" 입력단자 해석 + 전면 USB-C 완화 확인. 미해소 시 투찰 포기", "bad"),
        ("R2", "경영상태 배점 미달", "신용등급 A 이하 또는 등급확인서 미제출", "치명", None,
         "신용평가등급확인서 사전 발급(공고일 전일 이전 평가·유효기간 내). 1점 차로 S4가 탈락", "bad"),
        ("R3", "ASRock B860M-HDVP 600개 미확보", "에즈윈 재고·납기 부족", "높음", None,
         "(주)에즈윈 02-701-6350 사전 확약. 국내 대체 보드 없음 → 미확보 시 R1로 전이", "bad"),
        ("R4", "DDR5 16GB 600개 조달가 상승", "D램 공급부족 지속 (계약기간 1년)", "높음", 87960000,
         "총판 물량 선확보 + 계약기간 단가 고정 협상. RAM 45만 시 -57,960,000 / 50만 시 -87,960,000", "warn"),
        ("R5", "실제 발주량이 예정수량 미달", "학사 수요 변동 (이의제기 불가)", "중간", 93211440,
         "현설에서 최소 발주량·발주 스케줄 확인. 발주 60% 시 마진 2.33억→1.40억", "warn"),
        ("R6", "예정수량 초과분 단가 고정", "초과 발주 + 부품 시세 상승", "중간", None,
         "현설에서 단가 조정 조항 협의. 불가 시 초과 리스크를 투찰가에 반영", "warn"),
        ("R7", "품목2(복합기) 역마진", "품목2 응찰 시 확정 발생", "확정", -2200000,
         "품목2 미응찰. 흑백복합기 예정단가 250,000 대비 조사단가 290,000", "warn"),
        ("R8", "34\" 모니터 마진 제로", "\"1 DP, 2 HDMI\" 규격 미완화", "중간", 17508600,
         "완화 시 크로스오버 34QW290GM(299,000)로 대당 +97,270, 180대 17,508,600 개선", "warn"),
        ("R9", "위약벌 + 손해배상 별도 청구", "계약 불이행·해지", "중간", None,
         "보증금 몰취로 종결되지 않음. 납기·물량 확약 후 투찰 [공고문 9-라, 14-나]", "warn"),
        ("R10", "이의제기 창구 부재", "평가 결과 불복", "중간", None,
         "결과 비공개·이의제기 불가. 소송 시 향후 대구대 입찰 영구 배제 → 리스크 수용 여부 사전 판단", "warn"),
        ("R11", "부대비용 원가 미반영", "조립·설치·1년 A/S·보증증권 수수료·제안서·출장", "중간", 30000000,
         "PC 대당 50,000원 반영 권고 (분석 시트 부대비용 셀)", "warn"),
        ("R12", "현장설명회 미참가", "8/26 15:00 불참", "치명", None,
         "입찰 자체가 불가. 일정 확정 + 대리인 위임장·재직증명서 준비", "bad"),
    ]
    for rid, name, cond, grade, amt, action, tone in risks:
        r = row_cells(ws, r, [
            ("A", rid, dict(bold=True, ha="center")),
            ("B", name, dict(bold=True)),
            ("C", cond, dict(size=9)),
            ("D", grade, dict(ha="center", bold=True, bg=bgmap[tone])),
            ("E", amt, dict(ha="right", fmt="#,##0")),
            ("F", action, dict(size=9)),
        ], height=32)
    r = note(ws, r, "영향(원)은 품목1 투찰 85% 시나리오 기준 추정치이며, 공란은 금액 산정 불가(정성 리스크)를 뜻한다.")

    # ------------------------------------------------------------------ 9
    r += 1
    r = section(ws, r, "9. 현장설명회(2026-08-26 15:00) 확인 질문")
    r = table_head(ws, r, [("A", "No"), ("B:D", "질문"), ("E:F", "목적 / 기대효과")])
    qs = [
        ("규격 판정", "물품규격 '부족' 판정의 단위는 무엇인가? I/O PORTS 행 전체가 1사양인가, 포트별 개별 사양인가?",
         "감점폭이 -5점과 -20점으로 갈린다. 당락을 직접 결정하는 질문."),
        ("규격 완화", "PC I/O '2 HDMI'를 메인보드 + 그래픽카드 합산 포트 수로 인정하는가?",
         "네이티브 HDMI 2포트 B860 보드가 존재하지 않음. 인정 시 S3 이상 가능."),
        ("규격 완화", "PC I/O '1 RGB(D-Sub)' 조건을 완화할 수 있는가? 국내 유통 B860 보드 중 D-Sub 탑재 제품이 1종뿐이다.",
         "미완화 시 단일 공급선(에즈윈) 의존. 완화 시 보드 선택폭 확대 + 원가 절감."),
        ("규격 완화", "PC 전면 'USB-C 1개'를 케이스 전면 베이 패널로 제공해도 인정되는가?",
         "인정 시 부족 사양 1개 해소 → 기술 48→53점."),
        ("규격 해석", "모니터 24\" 입력단자 '2 HDMI, D-Sub port'는 HDMI 2포트인가, HDMI+D-Sub 합쳐 2포트인가?",
         "엄격 해석 시 전조건 충족 신제품이 시장에 없음. 후자면 Dell E2425HSM(176,000)으로 해소."),
        ("규격 오기", "모니터 32\" 크기 '86.4cm'는 오타 아닌가? (32인치=81.3cm, 86.4cm는 34인치 값)",
         "제안 모델 적합성 판정 기준 확정."),
        ("규격 완화", "모니터 34\" 입력단자를 '1 DP, 2 HDMI' → '1 DP, 1 HDMI'로 완화 가능한가?",
         "완화 시 크로스오버 34QW290GM(299,000, 400nit/3000:1)로 대당 97,270원 절감 (180대 17,508,600원)."),
        ("규격 오기", "OS/브라우저 항목의 'Internet Explorer 11'은 삭제·대체되는가? (2022-06 지원종료)",
         "제안 불가 항목을 그대로 두면 전 업체가 부족 판정 대상이 된다."),
        ("평가", "입찰가격평가 40점의 산식은? (제안요청서에 '배점×'까지만 기재되어 있음)",
         "투찰율 결정에 직결. 최저가 기준 비례식 여부 확인."),
        ("평가", "물품규격 '우수' 판정 기준은 무엇인가? (5개 이상 우수 = 40점)",
         "원가를 얼마나 올려 규격을 상향할지 판단 근거."),
        ("계약", "예정수량의 최소 발주량이 보장되는가? 발주 스케줄은?",
         "수량 미달 리스크(R5) 정량화."),
        ("계약", "예정수량 초과 시 D램 등 부품 시세를 반영한 단가 조정 협의가 가능한가?",
         "초과 공급 의무(공고문 15-라)의 역마진 리스크(R6) 차단."),
        ("계약", "하자보수 1년의 기산점이 계약 종료일(2027-08-31)인가? (첫 납품분 기준 최대 2년)",
         "A/S 충당금 산정."),
        ("계약", "지체상금 요율은? (대구대학교 재무·회계규정 원문 비공개)",
         "납기 지연 리스크 정량화."),
    ]
    for i, (cat, q, why) in enumerate(qs, 1):
        r = row_cells(ws, r, [
            ("A", i, dict(ha="center", bold=True)),
            ("B:D", q, {}),
            ("E:F", why, dict(size=9, color=C_GRAY)),
        ], height=32)
        ws[f"A{r-1}"].comment = None

    # ------------------------------------------------------------------ 10
    r += 1
    r = section(ws, r, "10. 종합 판정")
    r = table_head(ws, r, [("A", "구분"), ("B:D", "판정"), ("E:F", "근거")])
    for item, verdict, basis, tone, h in [
        ("마진 판정", "성립. 품목1 단독·규격대응 구성(S3~S5 동일 BOM) 기준 손익분기 투찰율 69.6%. "
                   "투찰 85% 시 마진 233,028,600원(18.1%), 90% 시 308,923,600원(22.6%)",
         "PC 대당 원가 1,489,090원(2위 실주문가 + 부대비용 50,000) 기준. '분석' 시트는 원안(S0) 구성이라 값이 다름", "ok", 40),
        ("통과 판정", "불충분. 현 구성(S0)은 기술 33점 자동탈락이고, 보드 교체+그래픽카드로 갈 수 있는 "
                   "최고점도 43점(S3)으로 여전히 탈락이다. 48점 이상은 현설 완화가 있어야만 가능",
         "제안요청서 나-2) 자동탈락선 48점 / 시뮬레이션 S0~S6", "bad", 40),
        ("품목2", "미응찰 권고. 예정총액 36,500,000원 대비 조사원가 38,700,000원으로 확정 역마진",
         "흑백복합기 예정단가 250,000 vs 조사단가 290,000", "warn", 32),
        ("최종 권고", "조건부 참여. 선결 3건을 순서대로 확인하고, 하나라도 미충족이면 투찰하지 않는다: "
                   "① [최우선] 8/26 현설에서 규격 해석·완화 확보(질문 1~5번) — 이것 없이는 제품 구성을 어떻게 해도 자동탈락  "
                   "② (주)에즈윈 B860M-HDVP 600개 납기·단가 확약 (국내 대체 불가)  "
                   "③ 신용평가등급확인서 AA 이상 확보 (1점 차로 당락)",
         "R1·R2·R3가 모두 치명 등급이며 R3 미해결은 R1로 전이된다", "warn", 60),
    ]:
        r = row_cells(ws, r, [
            ("A", item, dict(bold=True, bg=C_LBL_BG, ha="center")),
            ("B:D", verdict, dict(bold=True)),
            ("E:F", basis, dict(size=9)),
        ], height=h, bg=bgmap[tone] if tone else None)

    r = note(ws, r, "이 시트의 수치는 계약 조건 검토용 별도 산정이며, 원가·마진의 공식 값은 '분석' 시트를 기준으로 한다.")
    r = note(ws, r, "작성: Aside / 2026-08-20. 출처 표기가 없는 항목은 추정이며 '추정'으로 명시했다.")

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    # 인쇄: 가로, 폭 1페이지 맞춤
    ws.print_area = f"A1:F{r}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    return wb, r


if __name__ == "__main__":
    path = sys.argv[1]
    wb, last = build(path)
    # 저장 직전 재읽기: 사용자 병행 편집 확인
    check = openpyxl.load_workbook(path)
    if check["분석"]["H42"].value != "=SUM(H26:H41)":
        print("WARN: '분석' 시트 합계 수식이 변경됨 →", check["분석"]["H42"].value)
    wb.save(path)
    print(f"OK: '{SHEET}' 시트 추가 완료 ({last-1}행) → {path}")
    print("시트 순서:", wb.sheetnames)
