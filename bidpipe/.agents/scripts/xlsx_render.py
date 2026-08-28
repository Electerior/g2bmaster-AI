#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""xlsx -> HTML 렌더러 (soffice/Excel 없이 워크북을 그림으로 보여주기 위한 경로)

이 맥의 샌드박스에서는 GUI 앱을 못 띄운다:
  - soffice --headless --convert-to pdf  -> Abort trap: 6
  - qlmanage -t                          -> sandbox initialization failed
  - open -a "Microsoft Excel" / AppleScript -> LaunchServices/AppleEvents 차단(-1728)
그래서 렌더링은 "openpyxl로 서식까지 읽어 HTML을 만들고, 브라우저(Playwright)로
스크린샷·PDF를 뜨는" 경로로 한다. 이 스크립트는 그중 HTML 생성까지를 담당한다.

사용법:
    python .agents/scripts/xlsx_render.py <파일.xlsx> [-o out.html] [--sheet 이름]
                                          [--max-rows N] [--max-cols N]
    # 출력 HTML 경로를 stdout에 찍는다.

브라우저 캡처(REPL 쪽):
    const p = await openTab('about:blank');
    const html = await fs.readFile(htmlPath, 'utf8');
    await p.evaluate(h => { document.open(); document.write(h); document.close(); }, html);
    const png = await p.screenshot({ fullPage: true, type: 'png' });   // 이미지
    const pdf = await p.pdf({ format: 'A4', printBackground: true });  // PDF
"""

import argparse
import datetime as _dt
import html as _html
import os
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- 색

# openpyxl이 theme 색을 rgb로 안 풀어주므로 표준 Office 테마 팔레트로 근사한다.
_THEME = [
    "FFFFFF", "000000", "E7E6E6", "44546A", "4472C4", "ED7D31",
    "A5A5A5", "FFC000", "5B9BD5", "70AD47", "0563C1", "954F72",
]


def _tint(hexrgb, tint):
    if not tint:
        return hexrgb
    r, g, b = (int(hexrgb[i:i + 2], 16) for i in (0, 2, 4))
    if tint < 0:
        f = 1 + tint
        r, g, b = int(r * f), int(g * f), int(b * f)
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    return "%02X%02X%02X" % (max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b)))


def _color(c):
    """openpyxl Color -> '#RRGGBB' 또는 None"""
    if c is None:
        return None
    try:
        if c.type == "rgb" and isinstance(c.rgb, str) and len(c.rgb) == 8:
            if c.rgb[:2] == "00":          # 완전 투명
                return None
            return "#" + c.rgb[2:]
        if c.type == "theme":
            idx = int(c.theme)
            base = _THEME[idx] if 0 <= idx < len(_THEME) else "000000"
            return "#" + _tint(base, getattr(c, "tint", 0) or 0)
        if c.type == "indexed":
            from openpyxl.styles.colors import COLOR_INDEX
            v = COLOR_INDEX[int(c.indexed)]
            if isinstance(v, str) and len(v) == 8 and v[:2] != "00":
                return "#" + v[2:]
    except Exception:
        return None
    return None


# ---------------------------------------------------------------- 값 포맷

def _fmt(value, number_format):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        if isinstance(value, _dt.datetime):
            if value.hour or value.minute or value.second:
                return value.strftime("%Y-%m-%d %H:%M")
            return value.strftime("%Y-%m-%d")
        if isinstance(value, _dt.date):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        nf = (number_format or "").lower()
        if "%" in nf:
            digits = 0
            if "0.0" in nf:
                digits = nf.split("0.")[1].count("0")
            return ("{:,.%df}%%" % digits).format(value * 100)
        if "#,##" in nf or "," in nf or "\\" in nf or "₩" in nf or "$" in nf:
            prefix = "₩" if ("₩" in (number_format or "") or "\\" in nf) else ""
            if isinstance(value, float) and abs(value - round(value)) > 1e-9:
                return prefix + "{:,.2f}".format(value)
            return prefix + "{:,.0f}".format(round(value))
        if isinstance(value, float):
            if abs(value - round(value)) < 1e-9:
                return str(int(round(value)))
            return ("%g" % value)
        return str(value)
    return str(value)


_BORDER_W = {"thin": "1px", "hair": "1px", "medium": "2px", "thick": "3px",
             "double": "3px", "dotted": "1px", "dashed": "1px"}


def _border_css(side):
    if side is None or side.style is None:
        return None
    style = "solid"
    if side.style in ("dotted",):
        style = "dotted"
    elif side.style in ("dashed", "mediumDashed", "dashDot", "mediumDashDot"):
        style = "dashed"
    elif side.style == "double":
        style = "double"
    w = _BORDER_W.get(side.style, "1px")
    col = _color(side.color) or "#9aa0a6"
    return "%s %s %s" % (w, style, col)


# ---------------------------------------------------------------- 렌더

def render_sheet(ws, max_rows=None, max_cols=None):
    dim_max_row = ws.max_row or 1
    dim_max_col = ws.max_column or 1
    if max_rows:
        dim_max_row = min(dim_max_row, max_rows)
    if max_cols:
        dim_max_col = min(dim_max_col, max_cols)

    # 병합 셀 맵
    span = {}       # (r,c) -> (rowspan, colspan)
    covered = set()  # 병합에 먹힌 셀
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        span[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    covered.add((r, c))

    out = []
    out.append('<table class="sheet"><colgroup>')
    for c in range(1, dim_max_col + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        width = getattr(dim, "width", None) if dim else None
        px = int(round((width or 8.43) * 7.2 + 5))
        out.append('<col style="width:%dpx">' % px)
    out.append("</colgroup><tbody>")

    for r in range(1, dim_max_row + 1):
        rd = ws.row_dimensions.get(r)
        h = getattr(rd, "height", None) if rd else None
        style_row = ' style="height:%dpx"' % int(round(h * 1.34)) if h else ""
        out.append("<tr%s>" % style_row)
        for c in range(1, dim_max_col + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            rs, cs = span.get((r, c), (1, 1))
            css = []

            f = cell.font
            if f is not None:
                if f.bold:
                    css.append("font-weight:700")
                if f.italic:
                    css.append("font-style:italic")
                if f.underline:
                    css.append("text-decoration:underline")
                if f.size:
                    css.append("font-size:%gpx" % (float(f.size) * 1.05))
                fc = _color(f.color)
                if fc and fc.upper() != "#000000":
                    css.append("color:%s" % fc)

            fill = cell.fill
            if fill is not None and fill.patternType == "solid":
                bg = _color(fill.fgColor)
                if bg and bg.upper() not in ("#FFFFFF",):
                    css.append("background:%s" % bg)

            al = cell.alignment
            if al is not None:
                if al.horizontal:
                    css.append("text-align:%s" % ("center" if al.horizontal == "center"
                                                  else "right" if al.horizontal == "right"
                                                  else "left"))
                elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    css.append("text-align:right")
                if al.vertical:
                    css.append("vertical-align:%s" % ("middle" if al.vertical == "center"
                                                      else "bottom" if al.vertical == "bottom"
                                                      else "top"))
                if al.wrap_text:
                    css.append("white-space:pre-wrap")
                if al.indent:
                    css.append("padding-left:%dpx" % (6 + int(al.indent) * 10))
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                css.append("text-align:right")

            b = cell.border
            if b is not None:
                for name, side in (("top", b.top), ("bottom", b.bottom),
                                   ("left", b.left), ("right", b.right)):
                    v = _border_css(side)
                    if v:
                        css.append("border-%s:%s" % (name, v))

            text = _fmt(cell.value, cell.number_format)
            body = _html.escape(text).replace("\n", "<br>")
            link = getattr(cell, "hyperlink", None)
            target = getattr(link, "target", None) if link else None
            if target:
                body = '<a href="%s" title="%s">%s</a>' % (
                    _html.escape(target, quote=True),
                    _html.escape(target, quote=True),
                    body or _html.escape(target[:60]))

            attrs = ""
            if rs > 1:
                attrs += ' rowspan="%d"' % rs
            if cs > 1:
                attrs += ' colspan="%d"' % cs
            if css:
                attrs += ' style="%s"' % ";".join(css)
            out.append("<td%s>%s</td>" % (attrs, body))
        out.append("</tr>")
    out.append("</tbody></table>")
    return "".join(out)


_CSS = """
:root { color-scheme: light; }
body { margin:0; padding:18px 20px; background:#fff;
       font-family:-apple-system,'Apple SD Gothic Neo','Malgun Gothic',sans-serif;
       color:#111; }
h2.file { font-size:15px; margin:0 0 2px; }
div.path { font-size:11px; color:#777; margin:0 0 14px; word-break:break-all; }
h3.tab { font-size:13px; margin:18px 0 6px; padding:3px 8px; background:#eef2f7;
         border-left:3px solid #1f3864; display:inline-block; }
table.sheet { border-collapse:collapse; table-layout:fixed; font-size:12px;
              border:1px solid #d0d5dd; }
table.sheet td { border:1px solid #e3e6ea; padding:3px 6px; vertical-align:top;
                 overflow-wrap:anywhere; }
table.sheet a { color:#1155cc; }
"""


def render_workbook(path, sheet=None, max_rows=None, max_cols=None):
    wb = load_workbook(path, data_only=True)
    names = [sheet] if sheet else wb.sheetnames
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'><style>%s</style></head><body>" % _CSS,
        "<h2 class='file'>%s</h2>" % _html.escape(os.path.basename(path)),
        "<div class='path'>%s</div>" % _html.escape(path),
    ]
    for name in names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        if ws.sheet_state != "visible":
            continue
        parts.append("<h3 class='tab'>%s</h3>" % _html.escape(name))
        parts.append(render_sheet(ws, max_rows=max_rows, max_cols=max_cols))
    parts.append("</body></html>")
    return "".join(parts)


def main():
    ap = argparse.ArgumentParser(description="xlsx -> HTML (soffice 없이 렌더)")
    ap.add_argument("xlsx")
    ap.add_argument("-o", "--out", default=None, help="출력 HTML 경로 (기본: 입력과 같은 이름 .html)")
    ap.add_argument("--sheet", default=None, help="특정 시트만")
    ap.add_argument("--max-rows", type=int, default=None)
    ap.add_argument("--max-cols", type=int, default=None)
    args = ap.parse_args()

    src = os.path.abspath(args.xlsx)
    if not os.path.exists(src):
        print("파일 없음: %s" % src, file=sys.stderr)
        return 2
    html = render_workbook(src, sheet=args.sheet,
                           max_rows=args.max_rows, max_cols=args.max_cols)
    out = args.out or (os.path.splitext(src)[0] + ".html")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write(html)
    print(out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
