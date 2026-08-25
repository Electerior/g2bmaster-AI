# -*- coding: utf-8 -*-
"""분석 xlsx의 URL 열(E26:E45)을 진짜 하이퍼링크로 바꾸고 불량 URL을 감사한다.

사용법:
    python fix_links.py 20260819/*.xlsx        # 하이퍼링크 적용 + 감사
    python fix_links.py --audit 20260819/*.xlsx  # 감사만 (파일 수정 안 함)

2026-08-19 사용자 지적: 엑셀 URL이 텍스트로만 들어가 클릭이 안 되고,
일부는 다나와 검색결과/홈으로 걸려 상품에 바로 도달하지 못했음.
"""
import sys, os, glob
from openpyxl import load_workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gen_analysis import url_problem  # noqa: E402

LINK_FONT = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
ROW_FROM, ROW_TO, URL_COL = 26, 45, 5   # E26:E45


def process(path, audit_only=False):
    wb = load_workbook(path)
    ws = wb["분석"] if "분석" in wb.sheetnames else wb[wb.sheetnames[0]]
    linked, bad = 0, []
    for r in range(ROW_FROM, ROW_TO + 1):
        cell = ws.cell(row=r, column=URL_COL)
        u = (cell.value or "")
        if not isinstance(u, str) or not u.strip():
            continue
        u = u.strip()
        sel = str(ws.cell(row=r, column=7).value or "").strip().upper()
        prob = url_problem(u)
        if prob:
            bad.append((r, sel, ws.cell(row=r, column=1).value, u, prob))
        if u.lower().startswith("http") and not audit_only:
            if cell.hyperlink is None or getattr(cell.hyperlink, "target", None) != u:
                cell.hyperlink = u
                linked += 1
            cell.font = LINK_FONT
    if not audit_only and linked:
        wb.save(path)
    return linked, bad


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--audit"]
    audit_only = "--audit" in sys.argv
    files = []
    for a in args:
        files.extend(sorted(glob.glob(a)))
    tot_link, tot_bad = 0, 0
    for f in files:
        if os.path.basename(f).startswith("~$"):
            continue
        linked, bad = process(f, audit_only)
        tot_link += linked
        tot_bad += len(bad)
        if linked or bad:
            print(f"\n■ {os.path.basename(f)}  (하이퍼링크 {linked}건 적용)")
            for r, sel, name, u, prob in bad:
                mark = "★선택O" if sel == "O" else "  선택X"
                print(f"   {mark} r{r} [{name}] {prob}\n           {u[:110]}")
    print(f"\n총 하이퍼링크 {tot_link}건 적용, 불량 URL {tot_bad}건")
