# -*- coding: utf-8 -*-
"""price-db.json(플랫 배열) → products.json + observations.jsonl 일회성 마이그레이션.

하는 일
  1. 기존 price-db 엔트리를 '상품 정체성'과 '관측 1건'으로 쪼갠다
  2. 무URL 항목(당사 매입가표·견적)에 채널:모델코드 키를 부여한다
  3. 기존 observations.jsonl(구 형식)을 새 필드로 옮긴다 — 관측은 하나도 버리지 않는다
  4. 이름에 박힌 수량(`RAM x60`)을 떼고 qty_basis로 넘긴다
  5. price-db.json을 파생 뷰로 재생성한다

사용법
  python migrate_price_db.py --dry     # 결과만 출력
  python migrate_price_db.py           # 실제 기록 (원본은 .bak-<날짜>로 백업)
"""
import os, re, sys, json, glob, shutil, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from price_schema import (ROOT, DATA, PRODUCTS, OBS, VIEW, product_key, synthetic_key,
                          extract_model_code, infer_tier, save_products, rebuild_view)

DRY = "--dry" in sys.argv
STAMP = datetime.date.today().strftime("%Y%m%d")

QTY_IN_NAME = re.compile(r"\s*[x×]\s?\d+\s*$")
BRANDS = ["삼성전자", "삼성", "LG전자", "LG", "레노버", "Lenovo", "Dell", "델", "HP", "Apple",
          "애플", "ASUS", "에이수스", "MSI", "기가바이트", "GIGABYTE", "인텔", "Intel",
          "AMD", "엔비디아", "NVIDIA", "Micron", "마이크론", "Broadcom", "Seagate", "WD",
          "ABKO", "앱코", "알파스캔", "크로스오버", "TP-Link", "Supermicro", "ZOTAC"]


def clean_name(name):
    return QTY_IN_NAME.sub("", (name or "").strip()).strip() or "(무명)"


def guess_brand(*texts):
    for t in texts:
        for b in BRANDS:
            if b in (t or ""):
                return b
    return None


def workbook_names():
    """워크북에서 상품키 → 행 이름을 모은다 (obs에만 있는 키의 이름 보강용)."""
    from openpyxl import load_workbook
    out = {}
    for f in sorted(glob.glob(os.path.join(ROOT, "2026*/*.xlsx"))):
        if os.path.basename(f).startswith("~$"):
            continue
        try:
            wb = load_workbook(f)
        except Exception:
            continue
        ws = wb["분석"] if "분석" in wb.sheetnames else wb.active
        for r in range(26, 46):
            nm, u, sp = ws["A%d" % r].value, ws["E%d" % r].value, ws["B%d" % r].value
            k = product_key(str(u or ""))
            if k and nm and k not in out:
                out[k] = {"name": clean_name(str(nm)), "spec": str(sp or ""),
                          "url": str(u).strip(), "source_file": os.path.basename(f)}
    return out


def main():
    old = json.load(open(VIEW, encoding="utf-8"))
    if isinstance(old, dict):
        print("이미 새 형식(뷰)입니다. 중단."); return
    print("기존 price-db 엔트리: %d건" % len(old))

    products, obs_records = {}, []
    tier_count = {}

    for e in old:
        url = (e.get("url") or "").strip()
        name = clean_name(e.get("name"))
        spec = e.get("spec") or ""
        vendor = e.get("vendor") or ""
        key = product_key(url) or synthetic_key(vendor, name, spec)
        tier = infer_tier(vendor, url)
        tier_count[tier] = tier_count.get(tier, 0) + 1

        p = products.get(key)
        if not p:
            products[key] = p = {
                "name": name, "brand": guess_brand(name, spec),
                "model_code": extract_model_code(spec, name), "spec": spec,
                "qty_basis": e.get("qty_basis", "개당"),
                "volatility": e.get("volatility", "normal"),
                "urls": [url] if url.lower().startswith("http") else [],
                "aliases": [], "note": e.get("note", ""),
                "first_seen": e.get("date", ""), "source_file": e.get("source_file", ""),
            }
        else:
            # 중복 키: 더 서술적인 이름/스펙을 남기고 나머지는 alias로 보존
            if name != p["name"] and name not in p["aliases"]:
                p["aliases"].append(name)
                if len(name) > len(p["name"]):
                    p["name"], p["aliases"][-1] = name, p["name"]
            if len(spec) > len(p.get("spec") or ""):
                p["spec"] = spec
            if url.lower().startswith("http") and url not in p["urls"]:
                p["urls"].append(url)
            if e.get("note") and e["note"] not in (p.get("note") or ""):
                p["note"] = ((p.get("note") or "") + " | " + e["note"]).strip(" |")

        d = e.get("date") or STAMP
        obs_records.append({
            "key": key, "observed_at": "%sT00:00:00" % d,
            "observer": "ai:legacy-import", "method": "legacy", "tier": tier,
            "price": e.get("price"), "currency": "KRW",
            "fx_rate": None, "vat_included": e.get("vat_included", True),
            "vendor": vendor or None,
            "vendor_scope": ("internal" if "당사 매입" in vendor else
                             "lowest" if "최저가" in vendor else
                             "named" if vendor else None),
            "listing_status": "active", "genuine": None,
            "url": url or None, "note": e.get("note", ""),
            "source_file": e.get("source_file", ""),
        })

    # ── 기존 관측 로그(구 형식) 이관 ──
    wbn = workbook_names()
    migrated, new_keys = 0, 0
    if os.path.exists(OBS):
        for line in open(OBS, encoding="utf-8"):
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except ValueError:
                continue
            if r.get("observed_at"):        # 이미 새 형식
                obs_records.append(r); continue
            url = r.get("url") or ""
            # 구 로그의 key는 이전 버전 product_key로 만들어졌다. URL이 있으면 현행 규칙으로 재계산해야
            # 같은 상품이 두 개의 키로 갈라지지 않는다
            key = product_key(url) or r.get("key")
            if not key:
                continue
            vendor = r.get("vendor")
            obs_records.append({
                "key": key, "observed_at": r.get("scraped_at") or "%sT00:00:00" % STAMP,
                "observer": "script:price_refresh",
                "method": r.get("method") or "fetch",
                "tier": infer_tier(vendor, url, r.get("method")),
                "price": r.get("price"), "currency": "KRW", "fx_rate": None,
                "vat_included": True, "vendor": vendor,
                "vendor_scope": ("lowest" if not vendor else "named"),
                "listing_status": ("delisted" if r.get("delisted") else "active"),
                "genuine": None, "url": url or None,
                "core": r.get("core"), "delivery": r.get("delivery"),
                "error": r.get("error"),
            })
            migrated += 1
            if key not in products:
                meta = wbn.get(key, {})
                products[key] = {
                    "name": meta.get("name") or key, "brand": guess_brand(meta.get("name")),
                    "model_code": extract_model_code(meta.get("spec"), meta.get("name")),
                    "spec": meta.get("spec", ""), "qty_basis": "개당", "volatility": "normal",
                    "urls": [url] if url.lower().startswith("http") else [],
                    "aliases": [], "note": "관측 로그에서 발견 (price-db 미등록이던 상품)",
                    "first_seen": (r.get("scraped_at") or "")[:10],
                    "source_file": meta.get("source_file", ""),
                }
                new_keys += 1

    obs_records.sort(key=lambda r: (r.get("observed_at") or "", r.get("key") or ""))

    print("→ 상품(products) %d개 / 관측(observations) %d건" % (len(products), len(obs_records)))
    print("   구 관측 로그 이관 %d건, 그중 price-db에 없던 신규 상품 %d개" % (migrated, new_keys))
    print("   티어 분포(레거시 추론):", ", ".join("%s=%d" % kv for kv in sorted(tier_count.items())))
    synth = [k for k in products if k.split(":")[0] in ("internal", "quote", "unlinked")]
    for pre in ("internal", "quote", "unlinked"):
        sel = [k for k in synth if k.startswith(pre + ":")]
        if sel:
            print("   %-9s %3d개  예: %s" % (pre, len(sel), sel[0]))

    if DRY:
        print("\n(dry-run, 미기록)"); return

    for f in (VIEW, OBS):
        if os.path.exists(f):
            shutil.copy(f, "%s.bak-%s" % (f, STAMP))
    save_products(products)
    with open(OBS, "w", encoding="utf-8") as fh:
        for r in obs_records:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")
    rows = rebuild_view()
    print("\n기록 완료")
    print("  %s  (%d 상품)" % (PRODUCTS, len(products)))
    print("  %s  (%d 관측)" % (OBS, len(obs_records)))
    print("  %s  (%d 행, 파생 뷰)" % (VIEW, len(rows)))
    print("  백업: *.bak-%s" % STAMP)


if __name__ == "__main__":
    main()
