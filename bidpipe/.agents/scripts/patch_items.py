# -*- coding: utf-8 -*-
"""분석 워크북의 물품내역(26~45행)을 부분 수정/분해하는 패처.

gen_analysis.py는 Format.xlsx에서 새로 만들기 때문에 meta를 다시 넣어야 한다.
이 스크립트는 **기존 파일의 물품내역 블록만** 갈아끼운다 (공고정보·개요·부대비용은 보존).

입력 JSON:
[{
  "file": "20260820/xxx.xlsx",
  "edits":  {"26": {"price": 123, "url": "...", "vendor": "...", "note": "...", "name": "...", "qty": 2}},
  "expand": {"32": [ {name,spec,qty,vendor,url,price,note}, ... ]},   # 한 행을 여러 행으로 분해
  "extra_cost": 300000            # 선택
}]
"""
import sys, json, io, os
from openpyxl import load_workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gen_analysis import url_problem  # URL 품질 검증 재사용

LINK_FONT = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
BASE = os.environ.get("BIDPIPE_ROOT") or os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
R0, R1 = 26, 45


def _layout(ws):
    """물품행 끝행·부대비용 행을 라벨로 찾는다 (audit_prices._layout와 같은 규칙).

    26~45 / B50 고정 가정은 두 번 깨졌다:
      1) gen_analysis가 20행 초과 공고에서 행을 늘릴 때 (2026-08-21)
      2) 사용자가 엑셀에서 물품행을 지워 아래 블록이 위로 밀릴 때 (2026-08-25,
         KAIST R26BK01690151에서 규격미달 2행 삭제 → B50이 '=B16' 문자열이 되어
         old_cost 계산이 TypeError로 죽었다)
    행 번호를 고정으로 잡지 말 것. 이 파일은 사용자와 병행 편집된다.
    """
    end = R1
    for r in range(R0, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == "상품가격 합계":
            end = r - 1
            break
    extra_row = None
    for r in range(end + 1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().startswith("부대비용"):
            extra_row = r
            break
    return end, extra_row


def _num(v):
    """엑셀 수기 편집으로 들어온 공백·문자열을 0으로 내린다."""
    return v if isinstance(v, (int, float)) else 0


def read_items(ws, end=None):
    items = []
    for r in range(R0, (end or R1) + 1):
        nm = ws.cell(row=r, column=1).value
        if not nm:
            continue
        items.append(dict(
            row=r, name=nm,
            spec=ws.cell(row=r, column=2).value or "",
            qty=ws.cell(row=r, column=3).value or 1,
            vendor=ws.cell(row=r, column=4).value or "",
            url=ws.cell(row=r, column=5).value or "",
            price=ws.cell(row=r, column=6).value or 0,
            sel=ws.cell(row=r, column=7).value or "X",
            note=ws.cell(row=r, column=9).value or "",
        ))
    return items


def write_items(ws, items, end=None):
    R1 = end or globals()["R1"]
    for r in range(R0, R1 + 1):                     # 블록 비우기 (H 수식은 유지)
        for c in (1, 2, 3, 4, 5, 6, 7, 9):
            ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=5).hyperlink = None
    r = R0
    for it in items:
        if r > R1:
            print(f"  WARN: 45행 초과 - '{it['name'][:30]}' 이하 누락", file=sys.stderr)
            break
        ws.cell(row=r, column=1).value = it["name"]
        ws.cell(row=r, column=2).value = it.get("spec", "")
        ws.cell(row=r, column=3).value = it.get("qty", 1)
        ws.cell(row=r, column=4).value = it.get("vendor", "")
        u = (it.get("url") or "").strip()
        cell = ws.cell(row=r, column=5)
        cell.value = u
        if u.lower().startswith("http"):
            cell.hyperlink = u
            cell.font = LINK_FONT
        ws.cell(row=r, column=6).value = it.get("price", 0)
        ws.cell(row=r, column=7).value = it.get("sel", "O")
        ws.cell(row=r, column=8).value = f'=IF($G{r}="O",$C{r}*$F{r},0)'
        ws.cell(row=r, column=9).value = it.get("note", "")
        r += 1
    return r - R0


def main(path):
    specs = json.load(io.open(path, encoding="utf-8"))
    report = []
    for s in specs:
        f = os.path.join(BASE, s["file"])
        wb = load_workbook(f)
        ws = wb["분석"]
        end_row, extra_row = _layout(ws)
        before = read_items(ws, end_row)
        base = _num(ws["B14"].value)
        rate = _num(ws["B15"].value)
        extra = _num(ws.cell(row=extra_row, column=2).value) if extra_row else 0
        old_cost = sum(_num(i["qty"] or 1) * _num(i["price"]) for i in before if i["sel"] == "O") + extra

        edits = {int(k): v for k, v in (s.get("edits") or {}).items()}
        expand = {int(k): v for k, v in (s.get("expand") or {}).items()}
        out = []
        for it in before:
            if it["row"] in expand:
                for nu in expand[it["row"]]:
                    nu.setdefault("sel", "O")
                    out.append(nu)
                continue
            if it["row"] in edits:
                it.update(edits[it["row"]])
            out.append(it)
        if len(out) > end_row - R0 + 1:
            print(f"  ERROR {s['file']}: 물품행 {len(out)}개가 블록({R0}~{end_row})을 넘는다 - 중단",
                  file=sys.stderr)
            continue
        write_items(ws, out, end_row)
        if "extra_cost" in s and extra_row:
            ws.cell(row=extra_row, column=2).value = s["extra_cost"]
        for it in out:                              # URL 품질 검증
            if it.get("sel") == "O" and not it.get("quote"):
                p = url_problem(it.get("url"))
                if p:
                    print(f"  WARN {s['file']} [{it['name'][:34]}] {p}", file=sys.stderr)
        new_extra = _num(ws.cell(row=extra_row, column=2).value) if extra_row else 0
        new_cost = sum(_num(i.get("qty") or 1) * _num(i.get("price")) for i in out if i.get("sel") == "O") + new_extra
        wb.save(f)
        bid = round(base * rate)
        report.append(dict(file=os.path.basename(f), bid=bid,
                           old_cost=old_cost, new_cost=new_cost,
                           old_margin=round((bid - old_cost) / bid * 100, 1) if bid else None,
                           new_margin=round((bid - new_cost) / bid * 100, 1) if bid else None,
                           rows=len(out)))
    print(json.dumps(report, ensure_ascii=False, indent=1))


if __name__ == "__main__":
    main(sys.argv[1])
