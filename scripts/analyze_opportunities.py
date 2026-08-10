#!/usr/bin/env python3
"""KBID 사업기회분석 워크북을 우리 엔진으로 분류하고, 기존 분류와 대조한다.

```
python scripts/analyze_opportunities.py test/KBID_0806_사업기회분석.xlsx
python scripts/analyze_opportunities.py <xlsx> --baseline out/v1.json   # 이식 전후 비교
```

워크북에는 사람이 배정한 자회사·제품·Fit 이 이미 들어 있다. 그것을 정답 삼아 대조하므로
"돌아갔다"가 아니라 **"얼마나 맞았다"**를 말할 수 있다. 실측 이력은 `docs/opportunity-eval.md`.

엑셀로 내보낼 때 **우리 엔진이 채운 칸은 색과 🤖 로 표시한다.** 사람이 만든 분류와 기계가
만든 분류가 같은 표에서 구분되지 않으면, 다음에 그 파일을 여는 사람은 둘을 합쳐 읽는다.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
import time
from collections import Counter
from dataclasses import asdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.opportunity import Notice, classify, prompt_version  # noqa: E402

SUBS = ("서버스테이션", "브레인웨어", "일렉테리어", "렌탈본점")

# 워크북 '전체마스터' 시트의 열 순서.
COLUMNS = [
    "우선순위", "긴급도", "자회사", "Fit", "제품", "공고명", "공고기관", "수요기관",
    "D-day", "투찰마감일시", "구분", "지역제한", "금액(억원)", "투찰하한율",
]


def load_workbook_rows(path: Path) -> list[dict]:
    import openpyxl

    ws = openpyxl.load_workbook(path, read_only=True)["전체마스터"]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(COLUMNS, values))
        if not record.get("공고명"):
            continue
        record["자회사"] = "".join(ch for ch in str(record.get("자회사") or "") if ch.isalnum())
        rows.append(record)
    return rows


def to_notices(rows: list[dict]) -> list[Notice]:
    return [
        Notice(
            name=str(row["공고명"]).strip(),
            institution=str(row.get("공고기관") or "").strip(),
            amount=row.get("금액(억원)"),
            method=str(row.get("구분") or "").strip(),
            region=str(row.get("지역제한") or "").strip(),
        )
        for row in rows
    ]


def verdict_of(row: dict) -> str:
    """워크북 분류와 우리 판정의 관계. 표에 그대로 실린다."""
    ai = row.get("_ai")
    if not ai:
        return "⚠️미분석"
    truth, pred = row["자회사"], ai["subsidiary"]
    if truth == "확인필요":
        return "🆕신규발굴" if pred in SUBS else "동일(미분류)"
    return "일치" if truth == pred else "불일치"


def score(rows: list[dict]) -> dict:
    covered = [r for r in rows if r.get("_ai")]
    labeled = [r for r in covered if r["자회사"] in SUBS]
    hit = [r for r in labeled if r["_ai"]["subsidiary"] == r["자회사"]]
    unknown = [r for r in covered if r["자회사"] == "확인필요"]
    recovered = [r for r in unknown if r["_ai"]["subsidiary"] in SUBS]
    return {
        "공고": len(rows),
        "분석완료": len(covered),
        "미분석": len(rows) - len(covered),
        "배정건": len(labeled),
        "일치": len(hit),
        "일치율": round(len(hit) / len(labeled) * 100, 1) if labeled else 0.0,
        "확인필요": len(unknown),
        "신규발굴": len(recovered),
        "신규발굴률": round(len(recovered) / len(unknown) * 100, 1) if unknown else 0.0,
        "fit_High": sum(1 for r in covered if r["_ai"]["fit"] == "High"),
        "fit_High_워크북": sum(1 for r in rows if r.get("Fit") == "High"),
    }


# ── 엑셀 ─────────────────────────────────────────────────────────────────────
def write_xlsx(rows: list[dict], out: Path, meta: dict, compare: dict | None) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HUMAN = PatternFill("solid", fgColor="E8EAED")   # 워크북(사람) 분류
    ENGINE = PatternFill("solid", fgColor="D6E9F8")  # 우리 엔진이 채운 칸
    ENGINE_HEAD = PatternFill("solid", fgColor="1A73E8")
    TITLE = Font(bold=True, size=13)

    wb = openpyxl.Workbook()

    # ── 요약 ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "요약"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 20

    line = 1

    def put(label: str, value=None, bold: bool = False) -> None:
        nonlocal line
        cell = ws.cell(row=line, column=1, value=label)
        if bold:
            cell.font = Font(bold=True)
        if value is not None:
            ws.cell(row=line, column=2, value=value)
        line += 1

    ws.cell(row=1, column=1, value="KBID 사업기회분석 — AI 분류 결과").font = TITLE
    line = 3
    put("분석 엔진", "g2bmaster-AI · app/opportunity.py", bold=True)
    put("모델", meta["model"])
    put("프롬프트 버전", meta["prompt_version"])
    put("실행 시각", meta["ran_at"])
    put("소요", f"{meta['seconds']}초 · LLM 호출 {meta['calls']}회")
    put("입력 파일", meta["source"])
    line += 1

    put("표시 규칙", bold=True)
    cell = ws.cell(row=line, column=1, value="  🤖 붙은 열")
    cell.fill = ENGINE
    ws.cell(row=line, column=2, value="우리 엔진(LLM)이 채운 칸입니다. 사람이 검토하지 않았습니다.")
    line += 1
    cell = ws.cell(row=line, column=1, value="  그 외 열")
    cell.fill = HUMAN
    ws.cell(row=line, column=2, value="원본 워크북의 기존 분류입니다. 엔진은 건드리지 않았습니다.")
    line += 2

    put("측정", bold=True)
    stats = score(rows)
    labels = {
        "공고": "공고 건수",
        "분석완료": "🤖 분석 완료",
        "미분석": "🤖 미분석(응답 유실)",
        "배정건": "워크북이 자회사를 배정한 건",
        "일치": "  그중 판정 일치",
        "일치율": "  일치율(%)",
        "확인필요": "워크북 '확인필요' 건",
        "신규발굴": "  🤖 자회사를 지목한 건",
        "신규발굴률": "  신규발굴률(%)",
    }
    ws.cell(row=line, column=2, value="이번 실행").font = Font(bold=True)
    if compare:
        ws.cell(row=line, column=3, value="이식 전(baseline)").font = Font(bold=True)
    line += 1
    for key, label in labels.items():
        ws.cell(row=line, column=1, value=label)
        ws.cell(row=line, column=2, value=stats[key])
        if compare:
            before = compare.get(key)
            ws.cell(row=line, column=3, value=before)
        line += 1

    line += 1
    put("주의", bold=True)
    put("  ", "fit 은 등급 기준을 프롬프트에 적어 준 값이라 사람 판단과 다를 수 있습니다.")
    put("  ", f"워크북 High {stats['fit_High_워크북']}건 → 엔진 High {stats['fit_High']}건")

    # ── 전체분석 ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("전체분석")
    ai_headers = ["🤖업무구분", "🤖자회사", "🤖제품", "🤖fit", "🤖근거", "🤖판정"]
    headers = COLUMNS + ai_headers
    ws.append(headers)
    for index, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF" if name.startswith("🤖") else "000000")
        cell.fill = ENGINE_HEAD if name.startswith("🤖") else HUMAN
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ai = row.get("_ai") or {}
        ws.append([row.get(name) for name in COLUMNS] + [
            ai.get("work_kind", ""),
            ai.get("subsidiary", ""),
            ai.get("product", ""),
            ai.get("fit", ""),
            ai.get("reason", ""),
            verdict_of(row),
        ])

    first_ai = len(COLUMNS) + 1
    for excel_row in range(2, ws.max_row + 1):
        for column in range(first_ai, len(headers) + 1):
            ws.cell(row=excel_row, column=column).fill = ENGINE

    widths = {"공고명": 52, "공고기관": 26, "수요기관": 24, "🤖근거": 30, "🤖제품": 16}
    for index, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(name, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── 신규발굴 · 불일치 ───────────────────────────────────────────────────
    def sheet_for(title: str, picked: list[dict], note: str) -> None:
        sheet = wb.create_sheet(title)
        sheet.cell(row=1, column=1, value=note).font = Font(italic=True)
        head = ["금액(억원)", "공고명", "공고기관", "워크북 자회사", "🤖자회사", "🤖fit", "🤖근거"]
        sheet.append([])
        sheet.append(head)
        for index, name in enumerate(head, start=1):
            cell = sheet.cell(row=3, column=index)
            cell.font = Font(bold=True, color="FFFFFF" if name.startswith("🤖") else "000000")
            cell.fill = ENGINE_HEAD if name.startswith("🤖") else HUMAN
        for row in sorted(picked, key=lambda r: -(r.get("금액(억원)") or 0)):
            ai = row["_ai"]
            sheet.append([
                row.get("금액(억원)"), row["공고명"], row.get("공고기관"),
                row["자회사"], ai["subsidiary"], ai["fit"], ai["reason"],
            ])
        for excel_row in range(4, sheet.max_row + 1):
            for column in (5, 6, 7):
                sheet.cell(row=excel_row, column=column).fill = ENGINE
        for index, name in enumerate(head, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = {
                "공고명": 52, "공고기관": 26, "🤖근거": 32
            }.get(name, 14)
        sheet.freeze_panes = "A4"

    sheet_for(
        "신규발굴",
        [r for r in rows if verdict_of(r) == "🆕신규발굴"],
        "워크북이 '확인필요'로 남긴 건 중 엔진이 자회사를 지목한 것 — 사람 검토가 필요합니다.",
    )
    sheet_for(
        "불일치",
        [r for r in rows if verdict_of(r) == "불일치"],
        "워크북과 엔진의 판정이 다른 건 — 어느 쪽이 맞는지 확인이 필요합니다.",
    )

    wb.save(out)


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-dir", type=Path, default=ROOT / "test" / "out")
    parser.add_argument("--baseline", type=Path, help="이식 전 결과 json — 요약에 나란히 싣는다")
    parser.add_argument("--limit", type=int, default=0, help="앞에서 N 건만 (연습용)")
    args = parser.parse_args()

    rows = load_workbook_rows(args.workbook)
    if args.limit:
        rows = rows[: args.limit]
    notices = to_notices(rows)

    from app.llm.client import loaded_model

    model = await loaded_model()
    print(f"공고 {len(rows)}건 · 모델 {model} · 프롬프트 {prompt_version()}", flush=True)

    started = time.time()

    def progress(done: int, total: int) -> None:
        elapsed = time.time() - started
        print(f"  {done}/{total} ({elapsed:.0f}초 · 잔여 약 {elapsed / done * (total - done):.0f}초)", flush=True)

    verdicts, stats = await classify(notices, progress=progress)
    seconds = round(time.time() - started)

    for row, verdict in zip(rows, verdicts):
        row["_ai"] = asdict(verdict) if verdict else None

    args.out_dir.mkdir(parents=True, exist_ok=True)
    stem = args.workbook.stem
    json_path = args.out_dir / f"{stem}_AI분석.json"
    xlsx_path = args.out_dir / f"{stem}_AI분석.xlsx"

    meta = {
        "model": model,
        "prompt_version": prompt_version(),
        "ran_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "seconds": seconds,
        "calls": stats.calls,
        "source": args.workbook.name,
    }
    json_path.write_text(
        json.dumps({"meta": meta, "stats": asdict(stats), "rows": rows}, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )

    compare = None
    if args.baseline and args.baseline.exists():
        baseline = json.loads(args.baseline.read_text(encoding="utf-8"))
        compare = score(baseline["rows"] if isinstance(baseline, dict) else baseline)

    write_xlsx(rows, xlsx_path, meta, compare)

    result = score(rows)
    print(f"\n{'항목':<28}{'이번':>10}" + (f"{'이식 전':>12}" if compare else ""))
    for key in ("분석완료", "미분석", "일치", "일치율", "신규발굴", "신규발굴률", "fit_High"):
        before = f"{compare[key]:>12}" if compare else ""
        print(f"{key:<28}{result[key]:>10}{before}")
    print(f"\n재요청 {stats.retried}건 → 회수 {stats.recovered_by_retry}건 · "
          f"값집합 밖 폐기 {stats.rejected}건 · 실패 배치 {len(stats.failed_batches)}개")
    print(f"\n{json_path}\n{xlsx_path}")


if __name__ == "__main__":
    asyncio.run(main())
